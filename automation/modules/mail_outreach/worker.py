from __future__ import annotations

import csv
import hashlib
import json
import re
import urllib.parse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from playwright.async_api import TimeoutError as PlaywrightTimeoutError, async_playwright

from automation.adspower import AdsPowerClient
from automation.artifacts import setup_run_artifacts
from automation.browser import capture_screenshot, connect_profile
from automation.config import AdsPowerSettings
from automation.control_plane.storage import utcnow_iso
from automation.policies import load_farida_policy

from .state import AuditRecord, ContactRecord, MailOutreachState


STATUS_FILLS = {
    "ready_to_send": "D9EAF7",
    "thread_found_sent_only": "FFF2CC",
    "reply_unread": "C6EFCE",
    "reply_read_confirmed": "C6EFCE",
}


def _slug(value: str) -> str:
    return "".join(char if char.isalnum() or char in {"_", "-"} else "_" for char in value.strip().lower()) or "unknown"


def _sanitize_email(value: str) -> str:
    return str(value or "").strip().lower()


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def _save_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _mail_outreach_paths(project_root: Path) -> dict[str, Path]:
    base = project_root / "output" / "mail_outreach"
    return {
        "base": base,
        "registry_json": base / "contact_registry.json",
        "registry_md": base / "contact_registry.md",
        "registry_csv": base / "contact_registry.csv",
        "audit_json": base / "inbox_audit.json",
        "audit_md": base / "inbox_audit.md",
        "contacted_json": base / "contacted_contacts.json",
        "contacted_md": base / "contacted_contacts.md",
        "contacted_txt": base / "contacted_contacts.txt",
        "not_contacted_json": base / "not_contacted_contacts.json",
        "not_contacted_md": base / "not_contacted_contacts.md",
        "not_contacted_txt": base / "not_contacted_contacts.txt",
        "xlsx": base / "tables" / "mail_outreach_common.xlsx",
        "send_log": base / "send_log.jsonl",
    }


def _state_path(project_root: Path) -> Path:
    return project_root / "automation" / "state" / "mail_outreach_state.json"


def _manual_contacted_path(project_root: Path) -> Path:
    return project_root / "inputs" / "mail_outreach" / "already_contacted.txt"


def _load_manual_contacted(project_root: Path) -> dict[str, dict]:
    path = _manual_contacted_path(project_root)
    if not path.exists():
        return {}
    payload: dict[str, dict] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        parts = [item.strip() for item in (line.split("\t") if "\t" in line else line.split(maxsplit=2)) if item.strip()]
        if not parts:
            continue
        email = _sanitize_email(parts[0])
        if not email:
            continue
        payload[email] = {
            "email": email,
            "instagram_url": parts[1] if len(parts) >= 2 else "",
            "notes": parts[2] if len(parts) >= 3 else "",
        }
    return payload


def _append_manual_contacted_entry(
    project_root: Path,
    *,
    email: str,
    instagram_url: str = "",
    notes: str = "",
) -> None:
    normalized = _sanitize_email(email)
    if not normalized:
        return
    existing = _load_manual_contacted(project_root)
    if normalized in existing:
        current = existing[normalized]
        needs_write = False
        if instagram_url and not current.get("instagram_url"):
            current["instagram_url"] = instagram_url
            needs_write = True
        if notes and not current.get("notes"):
            current["notes"] = notes
            needs_write = True
        if not needs_write:
            return
    else:
        existing[normalized] = {
            "email": normalized,
            "instagram_url": instagram_url.strip(),
            "notes": notes.strip(),
        }
    path = _manual_contacted_path(project_root)
    lines = ["# email<TAB>instagram_url<TAB>optional notes"]
    for key in sorted(existing):
        item = existing[key]
        row = [item["email"]]
        if item.get("instagram_url") or item.get("notes"):
            row.append(item.get("instagram_url", ""))
        if item.get("notes"):
            while len(row) < 2:
                row.append("")
            row.append(item["notes"])
        lines.append("\t".join(row))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def _extract_brand_handle_from_instagram_url(url: str) -> str:
    text = str(url or "").strip()
    if not text:
        return ""
    match = re.search(r"instagram\.com/([^/?#]+)/?", text, re.IGNORECASE)
    if not match:
        return ""
    return match.group(1).strip().lstrip("@")


def _iter_brand_dossier_paths(project_root: Path) -> list[Path]:
    return sorted((project_root / "output" / "brand_intelligence").glob("*/brand_dossier.json"))


def _build_record_from_dossier(policy: dict, dossier_path: Path, email: str) -> ContactRecord:
    dossier = _load_json(dossier_path)
    instagram_profile = dict(dossier.get("instagram_profile") or {})
    arbiter_recommendation = dict(dossier.get("arbiter_recommendation") or {})
    ultra_threshold = int(policy.get("ultra_premium_policy", {}).get("follower_threshold") or 1_000_000)
    followers = int(instagram_profile.get("followers") or 0)
    return ContactRecord(
        contact_id=_slug(email),
        email=email,
        instagram_url=str(dossier.get("profile_url") or ""),
        brand_handle=str(dossier.get("brand_handle") or ""),
        brand_name=str(dossier.get("brand_name") or ""),
        source="dossier",
        source_details=["dossier_contact"],
        profile_url=str(dossier.get("profile_url") or ""),
        primary_site_url=str(dossier.get("primary_site_url") or ""),
        preferred_contact_channel="email",
        dossier_json_path=str(dossier_path),
        dossier_md_path=str(dossier_path.with_suffix(".md")),
        brand_folder=str(dossier_path.parent),
        brand_followers=followers,
        brand_posts=int(instagram_profile.get("posts") or 0),
        brand_value_tier=str(arbiter_recommendation.get("segment") or ""),
        is_ultra_premium_brand=followers >= ultra_threshold,
        special_handling=str(arbiter_recommendation.get("special_handling") or ""),
    )


def _build_record_from_manual_line(project_root: Path, policy: dict, line: str) -> ContactRecord | None:
    raw = line.strip()
    if not raw or raw.startswith("#"):
        return None
    parts = [item.strip() for item in (raw.split("\t") if "\t" in raw else raw.split(maxsplit=2)) if item.strip()]
    if not parts:
        return None
    email = _sanitize_email(parts[0])
    if "@" not in email:
        return None
    instagram_url = parts[1].strip() if len(parts) >= 2 else ""
    notes = parts[2].strip() if len(parts) >= 3 else ""
    brand_handle = _extract_brand_handle_from_instagram_url(instagram_url)
    dossier_path = project_root / "output" / "brand_intelligence" / _slug(brand_handle) / "brand_dossier.json"
    dossier = _load_json(dossier_path) if dossier_path.exists() else {}
    instagram_profile = dict(dossier.get("instagram_profile") or {})
    ultra_threshold = int(policy.get("ultra_premium_policy", {}).get("follower_threshold") or 1_000_000)
    followers = int(instagram_profile.get("followers") or 0)
    return ContactRecord(
        contact_id=_slug(email),
        email=email,
        instagram_url=instagram_url,
        brand_handle=brand_handle or str(dossier.get("brand_handle") or ""),
        brand_name=str(dossier.get("brand_name") or brand_handle or ""),
        source="manual_input",
        source_details=["manual_input"],
        notes=notes,
        profile_url=str(dossier.get("profile_url") or instagram_url),
        primary_site_url=str(dossier.get("primary_site_url") or ""),
        preferred_contact_channel="email",
        dossier_json_path=str(dossier_path) if dossier_path.exists() else "",
        dossier_md_path=str(dossier_path.with_suffix(".md")) if dossier_path.exists() else "",
        brand_folder=str(dossier_path.parent) if dossier_path.exists() else "",
        brand_followers=followers,
        brand_posts=int(instagram_profile.get("posts") or 0),
        is_ultra_premium_brand=followers >= ultra_threshold,
    )


def _build_record_from_already_contacted_line(project_root: Path, policy: dict, line: str) -> ContactRecord | None:
    raw = line.strip()
    if not raw or raw.startswith("#"):
        return None
    parts = [item.strip() for item in (raw.split("\t") if "\t" in raw else raw.split(maxsplit=2)) if item.strip()]
    if not parts:
        return None
    email = _sanitize_email(parts[0])
    if "@" not in email:
        return None
    instagram_url = parts[1].strip() if len(parts) >= 2 else ""
    notes = parts[2].strip() if len(parts) >= 3 else ""
    brand_handle = _extract_brand_handle_from_instagram_url(instagram_url)
    dossier_path = project_root / "output" / "brand_intelligence" / _slug(brand_handle) / "brand_dossier.json"
    dossier = _load_json(dossier_path) if dossier_path.exists() else {}
    instagram_profile = dict(dossier.get("instagram_profile") or {})
    ultra_threshold = int(policy.get("ultra_premium_policy", {}).get("follower_threshold") or 1_000_000)
    followers = int(instagram_profile.get("followers") or 0)
    return ContactRecord(
        contact_id=_slug(email),
        email=email,
        instagram_url=instagram_url,
        brand_handle=brand_handle or str(dossier.get("brand_handle") or ""),
        brand_name=str(dossier.get("brand_name") or brand_handle or ""),
        source="manual_already_contacted",
        source_details=["manual_already_contacted"],
        notes=notes,
        profile_url=str(dossier.get("profile_url") or instagram_url),
        primary_site_url=str(dossier.get("primary_site_url") or ""),
        preferred_contact_channel="email",
        dossier_json_path=str(dossier_path) if dossier_path.exists() else "",
        dossier_md_path=str(dossier_path.with_suffix(".md")) if dossier_path.exists() else "",
        brand_folder=str(dossier_path.parent) if dossier_path.exists() else "",
        brand_followers=followers,
        brand_posts=int(instagram_profile.get("posts") or 0),
        is_ultra_premium_brand=followers >= ultra_threshold,
    )


def _merge_contact_records(records: list[ContactRecord]) -> list[ContactRecord]:
    merged: dict[str, ContactRecord] = {}
    for record in records:
        key = _sanitize_email(record.email)
        if not key:
            continue
        if key not in merged:
            merged[key] = record
            continue
        current = merged[key]
        if not current.instagram_url and record.instagram_url:
            current.instagram_url = record.instagram_url
        if not current.brand_handle and record.brand_handle:
            current.brand_handle = record.brand_handle
        if not current.brand_name and record.brand_name:
            current.brand_name = record.brand_name
        if not current.profile_url and record.profile_url:
            current.profile_url = record.profile_url
        if not current.primary_site_url and record.primary_site_url:
            current.primary_site_url = record.primary_site_url
        if not current.dossier_json_path and record.dossier_json_path:
            current.dossier_json_path = record.dossier_json_path
            current.dossier_md_path = record.dossier_md_path
            current.brand_folder = record.brand_folder
        current.source = ",".join(sorted(set(filter(None, [*current.source.split(","), record.source]))))
        current.source_details = sorted(set([*current.source_details, *record.source_details]))
        if record.notes and record.notes not in current.notes:
            current.notes = "; ".join(part for part in [current.notes, record.notes] if part)
        current.brand_followers = max(current.brand_followers, record.brand_followers)
        current.brand_posts = max(current.brand_posts, record.brand_posts)
        current.is_ultra_premium_brand = current.is_ultra_premium_brand or record.is_ultra_premium_brand
        if not current.special_handling and record.special_handling:
            current.special_handling = record.special_handling
    return sorted(merged.values(), key=lambda item: (item.brand_handle or "", item.email))


def _write_markdown_table(path: Path, headers: list[str], rows: list[list[str]], title: str) -> None:
    lines = [f"# {title}", "", "| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        escaped = [str(cell or "").replace("\n", " ").replace("|", "\\|") for cell in row]
        lines.append("| " + " | ".join(escaped) + " |")
    lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def _write_csv(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def _set_header_style(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def _autosize_sheet(sheet) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value or "")) + 2)
    for index, width in widths.items():
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = min(max(width, 12), 48)


def _write_common_workbook(project_root: Path, registry: list[ContactRecord], audits: list[AuditRecord]) -> Path:
    paths = _mail_outreach_paths(project_root)
    try:
        from automation.modules.instagram_dm_outreach.reporting import load_instagram_dm_status_rows
    except Exception:
        dm_rows = []
    else:
        dm_rows = load_instagram_dm_status_rows(project_root)
    dm_by_handle = {row.handle: row for row in dm_rows if row.handle}

    workbook = Workbook()
    contacts_sheet = workbook.active
    contacts_sheet.title = "Contacts"
    contacts_sheet.append(
        [
            "brand_handle",
            "brand_name",
            "email",
            "instagram_url",
            "primary_site_url",
            "followers",
            "posts",
            "source",
            "preferred_channel",
            "ultra_premium",
            "instagram_dm_status",
            "instagram_dm_sends",
            "instagram_dm_last_sent",
            "notes",
        ]
    )
    for item in registry:
        dm_row = dm_by_handle.get(item.brand_handle)
        contacts_sheet.append(
            [
                item.brand_handle,
                item.brand_name,
                item.email,
                item.instagram_url,
                item.primary_site_url,
                item.brand_followers,
                item.brand_posts,
                item.source,
                item.preferred_contact_channel,
                "yes" if item.is_ultra_premium_brand else "no",
                dm_row.status if dm_row else "",
                dm_row.send_count if dm_row else 0,
                dm_row.last_sent_at_iso if dm_row else "",
                item.notes,
            ]
        )
    _set_header_style(contacts_sheet)
    _autosize_sheet(contacts_sheet)

    audit_sheet = workbook.create_sheet("Audit")
    audit_sheet.append(
        [
            "brand_handle",
            "brand_name",
            "email",
            "status",
            "result_count",
            "unread_count",
            "local_sent_count",
            "preferred_channel",
            "detail",
            "query_url",
        ]
    )
    for item in audits:
        audit_sheet.append(
            [
                item.brand_handle,
                item.brand_name,
                item.email,
                item.status,
                item.result_count,
                item.unread_count,
                item.local_sent_count,
                item.preferred_contact_channel,
                item.detail,
                item.query_url,
            ]
        )
        if item.status in STATUS_FILLS:
            fill = PatternFill(fill_type="solid", fgColor=STATUS_FILLS[item.status])
            for cell in audit_sheet[audit_sheet.max_row]:
                cell.fill = fill
    _set_header_style(audit_sheet)
    _autosize_sheet(audit_sheet)

    dm_sheet = workbook.create_sheet("InstagramDM")
    dm_sheet.append(
        [
            "handle",
            "status",
            "send_count",
            "reply_detected",
            "last_sent_at_iso",
            "target_url",
            "detail",
            "last_artifact_dir",
        ]
    )
    dm_fills = {
        "ready_to_send": "D9EAF7",
        "thread_found_sent_only": "FFF2CC",
        "sent_multiple": "FFF2CC",
        "reply_detected": "C6EFCE",
    }
    for row in dm_rows:
        dm_sheet.append(
            [
                row.handle,
                row.status,
                row.send_count,
                "yes" if row.reply_detected else "no",
                row.last_sent_at_iso,
                row.target_url,
                row.detail,
                row.last_artifact_dir,
            ]
        )
        fill_code = dm_fills.get(row.status)
        if fill_code:
            fill = PatternFill(fill_type="solid", fgColor=fill_code)
            for cell in dm_sheet[dm_sheet.max_row]:
                cell.fill = fill
    _set_header_style(dm_sheet)
    _autosize_sheet(dm_sheet)

    paths["xlsx"].parent.mkdir(parents=True, exist_ok=True)
    workbook.save(paths["xlsx"])
    return paths["xlsx"]


def _append_jsonl(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False) + "\n")


def _build_subject_variants(brand_name: str) -> list[str]:
    brand = brand_name.strip() or "бренд"
    return [
        f"Предложение по сотрудничеству с Farida Shirinova для {brand}",
        f"Рекламная интеграция: Farida Shirinova × {brand}",
        f"Сотрудничество с Farida Shirinova для {brand}",
    ]


def build_master_message(policy: dict, *, brand_name: str, channel: str, seed_value: str) -> tuple[str, str]:
    metrics = dict(policy.get("audience_metrics") or {})
    creator = dict(policy.get("creator") or {})
    cta_policy = dict(policy.get("cta_policy") or {})
    followers = int(metrics.get("followers") or 0)
    views_30d = int(metrics.get("reels_views_30d") or 0)
    women_share = metrics.get("women_share_percent") or 0
    age_range = str(metrics.get("core_age_range") or "").strip()
    geo = str(metrics.get("primary_geo") or "").strip()
    moscow_share = metrics.get("moscow_share_percent") or 0
    rec_share = str(metrics.get("recommendation_share_percent") or 0).replace(".", ",")
    creator_name = str(creator.get("name") or "Farida Shirinova")
    creator_handle = str(creator.get("handle") or "farida.shirinova").strip()
    creator_url = f"https://www.instagram.com/{creator_handle}/"
    seed = int(hashlib.md5(seed_value.encode("utf-8")).hexdigest(), 16)

    intro_variants = [
        "Предлагаю обсудить рекламную интеграцию, потому что вижу прямое совпадение между Вашим брендом и аудиторией блогера.",
        "Предлагаю обсудить сотрудничество, так как считаю, что бренд органично может встроиться в аудиторию блогера.",
        "Пишу с предложением рекламной интеграции, поскольку вижу сильный fit между брендом и аудиторией Farida.",
    ]
    cta_variants = [
        str(cta_policy.get("preferred") or "Готов направить актуальную статистику и 2–3 идеи интеграции под Ваш продукт."),
        "Если интересно, готов направить актуальную статистику и 2–3 идеи интеграции под Ваш продукт.",
        "При интересе готов отправить актуальную статистику и предложить 2–3 идеи интеграции под Ваш бренд.",
    ]
    subject = _build_subject_variants(brand_name)[seed % 3]
    intro = intro_variants[seed % len(intro_variants)]
    cta = cta_variants[seed % len(cta_variants)]

    body_lines = [
        "Здравствуйте.",
        f"Меня зовут Владислав, я представляю {creator_name}.",
        f"Профиль блогера: @{creator_handle}",
        f"Ссылка на профиль: {creator_url}",
        intro,
        "Коротко по профилю:",
        f"— {followers:,} подписчиков".replace(",", " "),
        f"— {views_30d / 1_000_000:.1f} млн просмотров за последние 30 дней".replace(".", ","),
        f"— {women_share}% аудитории — женщины",
        f"— основное ядро аудитории — {age_range}",
        f"— основная география — {geo} (Москва — {moscow_share}%)",
        f"— {rec_share}% просмотров приходят из рекомендаций",
        f"Для {brand_name.strip() or 'Вашего бренда'} это рабочий формат на узнаваемость и расширение касаний с женской аудиторией.",
        cta,
    ]
    if channel == "website_contact":
        body_lines.append("Если у Вас есть отдельный контакт для PR или partnerships, буду признателен за направление обращения.")
    elif channel == "instagram_dm":
        body_lines.append("Если удобнее, можете передать сообщение менеджеру или PR-отделу.")
    else:
        body_lines.append("Если письма по рекламе ведёт отдельный PR- или partnership-отдел, буду признателен за контакт или перенаправление.")
    body_lines.extend(["", "С уважением,", "Владислав"])
    return subject, "\n".join(body_lines)


def _write_brand_mail_summary(
    project_root: Path,
    contact: ContactRecord,
    *,
    audit: AuditRecord | None = None,
    sent_payload: dict | None = None,
) -> None:
    brand_folder = Path(contact.brand_folder) if contact.brand_folder else None
    if brand_folder is None or not brand_folder.exists():
        return
    payload = {
        "brand_handle": contact.brand_handle,
        "brand_name": contact.brand_name,
        "email": contact.email,
        "instagram_url": contact.instagram_url,
        "preferred_contact_channel": contact.preferred_contact_channel,
        "source": contact.source,
        "is_ultra_premium_brand": contact.is_ultra_premium_brand,
        "special_handling": contact.special_handling,
        "audit": audit.to_dict() if audit else {},
        "last_sent": sent_payload or {},
    }
    summary_json = brand_folder / "mail_outreach.json"
    summary_md = brand_folder / "mail_outreach.md"
    _save_json(summary_json, payload)

    lines = [
        f"# Mail Outreach @{contact.brand_handle or 'unknown'}",
        "",
        f"- Brand name: {contact.brand_name or 'unknown'}",
        f"- Email: {contact.email}",
        f"- Instagram URL: {contact.instagram_url or 'not found'}",
        f"- Preferred channel: {contact.preferred_contact_channel}",
        f"- Source: {contact.source}",
        f"- Ultra premium: {'yes' if contact.is_ultra_premium_brand else 'no'}",
        f"- Special handling: {contact.special_handling or 'none'}",
    ]
    if audit:
        lines.extend(
            [
                "",
                "## Audit",
                f"- Status: {audit.status}",
                f"- Detail: {audit.detail}",
                f"- Query URL: {audit.query_url}",
                f"- Result count: {audit.result_count}",
                f"- Unread count: {audit.unread_count}",
                f"- Local sent count: {audit.local_sent_count}",
                f"- Updated at: {audit.updated_at_iso}",
            ]
        )
    if sent_payload:
        lines.extend(
            [
                "",
                "## Last Send",
                f"- Subject: {sent_payload.get('subject', '')}",
                f"- Sent at: {sent_payload.get('sent_at_iso', '')}",
                f"- Log path: {sent_payload.get('log_path', '')}",
                f"- Screenshot before: {sent_payload.get('screenshot_before', '')}",
                f"- Screenshot after: {sent_payload.get('screenshot_after', '')}",
            ]
        )
    summary_md.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")

    dossier_json_path = Path(contact.dossier_json_path) if contact.dossier_json_path else None
    if dossier_json_path and dossier_json_path.exists():
        dossier = _load_json(dossier_json_path)
        dossier["mail_outreach"] = {
            "email": contact.email,
            "preferred_contact_channel": contact.preferred_contact_channel,
            "status": audit.status if audit else "",
            "detail": audit.detail if audit else "",
            "last_sent_at_iso": (sent_payload or {}).get("sent_at_iso", ""),
            "mail_summary_path": str(summary_json),
        }
        _save_json(dossier_json_path, dossier)


def _build_contact_status_lists(project_root: Path, registry: list[ContactRecord], audits: list[AuditRecord]) -> tuple[list[dict], list[dict]]:
    state = MailOutreachState.load(_state_path(project_root))
    manual_contacted = _load_manual_contacted(project_root)
    audit_by_email = {item.email: item for item in audits}
    contacted: list[dict] = []
    not_contacted: list[dict] = []

    for item in registry:
        email = _sanitize_email(item.email)
        local_sent_count = len(state.sent_messages.get(email, []))
        manual_entry = manual_contacted.get(email)
        audit = audit_by_email.get(email)
        contacted_reason = ""
        status = audit.status if audit else ""

        if manual_entry:
            contacted_reason = "manual_already_contacted"
        elif local_sent_count > 0:
            contacted_reason = "local_send_log"
        elif status in {"thread_found_sent_only", "reply_unread", "reply_read_confirmed"}:
            contacted_reason = f"audit:{status}"

        payload = {
            "brand_handle": item.brand_handle,
            "brand_name": item.brand_name,
            "email": email,
            "instagram_url": item.instagram_url or (manual_entry or {}).get("instagram_url", ""),
            "status": status or ("contacted" if contacted_reason else "not_contacted"),
            "contacted_reason": contacted_reason or "not_contacted",
            "local_sent_count": local_sent_count,
            "notes": item.notes or (manual_entry or {}).get("notes", ""),
            "preferred_contact_channel": item.preferred_contact_channel,
        }
        if contacted_reason:
            contacted.append(payload)
        else:
            not_contacted.append(payload)

    contacted.sort(key=lambda row: (row["brand_handle"], row["email"]))
    not_contacted.sort(key=lambda row: (row["brand_handle"], row["email"]))
    return contacted, not_contacted


def build_contact_registry(project_root: Path, *, input_path: Path | None = None) -> list[ContactRecord]:
    policy = load_farida_policy(project_root)
    manual_input_path = input_path or (project_root / "inputs" / "mail_outreach" / "contacts.txt")
    already_contacted_path = _manual_contacted_path(project_root)
    records: list[ContactRecord] = []

    for dossier_path in _iter_brand_dossier_paths(project_root):
        dossier = _load_json(dossier_path)
        for email in list((dossier.get("contact_signals") or {}).get("emails") or []):
            sanitized = _sanitize_email(email)
            if sanitized:
                records.append(_build_record_from_dossier(policy, dossier_path, sanitized))

    if manual_input_path.exists():
        for line in manual_input_path.read_text(encoding="utf-8").splitlines():
            record = _build_record_from_manual_line(project_root, policy, line)
            if record:
                records.append(record)

    if already_contacted_path.exists():
        for line in already_contacted_path.read_text(encoding="utf-8").splitlines():
            record = _build_record_from_already_contacted_line(project_root, policy, line)
            if record:
                records.append(record)

    merged = _merge_contact_records(records)
    state = MailOutreachState.load(_state_path(project_root))
    state.last_registry_built_at_iso = utcnow_iso()
    state.save(_state_path(project_root))
    return merged


def write_mail_outreach_outputs(project_root: Path, registry: list[ContactRecord], audits: list[AuditRecord]) -> dict[str, str]:
    paths = _mail_outreach_paths(project_root)
    contacted, not_contacted = _build_contact_status_lists(project_root, registry, audits)

    _save_json(paths["registry_json"], [item.to_dict() for item in registry])
    _write_markdown_table(
        paths["registry_md"],
        ["brand_handle", "brand_name", "email", "instagram_url", "followers", "source", "ultra_premium", "notes"],
        [
            [
                item.brand_handle,
                item.brand_name,
                item.email,
                item.instagram_url,
                str(item.brand_followers),
                item.source,
                "yes" if item.is_ultra_premium_brand else "no",
                item.notes,
            ]
            for item in registry
        ],
        "Mail Contact Registry",
    )
    _write_csv(
        paths["registry_csv"],
        ["brand_handle", "brand_name", "email", "instagram_url", "primary_site_url", "followers", "posts", "source", "preferred_channel", "ultra_premium", "notes"],
        [
            [
                item.brand_handle,
                item.brand_name,
                item.email,
                item.instagram_url,
                item.primary_site_url,
                str(item.brand_followers),
                str(item.brand_posts),
                item.source,
                item.preferred_contact_channel,
                "yes" if item.is_ultra_premium_brand else "no",
                item.notes,
            ]
            for item in registry
        ],
    )

    _save_json(paths["audit_json"], [item.to_dict() for item in audits])
    _write_markdown_table(
        paths["audit_md"],
        ["brand_handle", "brand_name", "email", "status", "unread_count", "local_sent_count", "detail"],
        [
            [item.brand_handle, item.brand_name, item.email, item.status, str(item.unread_count), str(item.local_sent_count), item.detail]
            for item in audits
        ],
        "Mail Inbox Audit",
    )
    _save_json(paths["contacted_json"], contacted)
    _save_json(paths["not_contacted_json"], not_contacted)
    _write_markdown_table(
        paths["contacted_md"],
        ["brand_handle", "brand_name", "email", "status", "contacted_reason", "local_sent_count", "notes"],
        [
            [
                item["brand_handle"],
                item["brand_name"],
                item["email"],
                item["status"],
                item["contacted_reason"],
                str(item["local_sent_count"]),
                item["notes"],
            ]
            for item in contacted
        ],
        "Contacted Contacts",
    )
    _write_markdown_table(
        paths["not_contacted_md"],
        ["brand_handle", "brand_name", "email", "status", "preferred_contact_channel", "notes"],
        [
            [
                item["brand_handle"],
                item["brand_name"],
                item["email"],
                item["status"],
                item["preferred_contact_channel"],
                item["notes"],
            ]
            for item in not_contacted
        ],
        "Not Contacted Contacts",
    )
    paths["contacted_txt"].write_text(
        "\n".join(
            f"{item['email']}\t{item['instagram_url']}\t{item['contacted_reason']}"
            for item in contacted
        )
        + ("\n" if contacted else ""),
        encoding="utf-8-sig",
    )
    paths["not_contacted_txt"].write_text(
        "\n".join(
            f"{item['email']}\t{item['instagram_url']}\t{item['preferred_contact_channel']}"
            for item in not_contacted
        )
        + ("\n" if not_contacted else ""),
        encoding="utf-8-sig",
    )
    xlsx_path = _write_common_workbook(project_root, registry, audits)
    return {
        "registry_json": str(paths["registry_json"]),
        "registry_md": str(paths["registry_md"]),
        "registry_csv": str(paths["registry_csv"]),
        "audit_json": str(paths["audit_json"]),
        "audit_md": str(paths["audit_md"]),
        "contacted_json": str(paths["contacted_json"]),
        "contacted_md": str(paths["contacted_md"]),
        "contacted_txt": str(paths["contacted_txt"]),
        "not_contacted_json": str(paths["not_contacted_json"]),
        "not_contacted_md": str(paths["not_contacted_md"]),
        "not_contacted_txt": str(paths["not_contacted_txt"]),
        "xlsx": str(xlsx_path),
    }


async def _open_mail_session(project_root: Path, *, label: str):
    artifacts, logger = setup_run_artifacts(project_root, label)
    settings = AdsPowerSettings.from_project_root(project_root)
    client = AdsPowerClient(
        AdsPowerSettings(
            base_url=settings.base_url,
            api_key=settings.api_key,
            profile_no="337",
        )
    )
    started = client.start_profile(profile_no="337", last_opened_tabs=False)
    playwright = await async_playwright().start()
    browser, context, page = await connect_profile(playwright, started.ws_puppeteer, logger)
    for extra_page in list(context.pages)[1:]:
        try:
            await extra_page.close()
        except Exception:
            logger.warning("Failed to close extra tab")
    await page.goto("https://mail.ru/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1200)
    await page.goto("https://e.mail.ru/inbox/", wait_until="domcontentloaded", timeout=60000)
    try:
        await page.wait_for_load_state("networkidle", timeout=8000)
    except PlaywrightTimeoutError:
        logger.info("Mail.ru inbox did not reach networkidle")
    return client, playwright, browser, page, artifacts, logger


async def _close_mail_session(client: AdsPowerClient, playwright, browser) -> None:
    try:
        await browser.close()
    except Exception:
        pass
    try:
        await playwright.stop()
    except Exception:
        pass
    try:
        client.stop_profile(profile_no="337")
    except Exception:
        pass


async def _probe_search_result(page, query: str) -> dict:
    await page.goto("https://e.mail.ru/search/?q_query=" + urllib.parse.quote(query), wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1800)
    await page.evaluate(
        """
async (target) => {
  const lower = String(target || "").toLowerCase();
  for (let i = 0; i < 8; i += 1) {
    const text = (document.body.innerText || "").toLowerCase();
    if (text.includes(lower)) {
      return;
    }
    window.scrollTo(0, document.body.scrollHeight);
    await new Promise((resolve) => setTimeout(resolve, 400));
  }
}""",
        query,
    )
    return await page.evaluate(
        """
() => {
  const text = document.body.innerText || "";
  const datasetRoot =
    document.querySelector(".dataset__items") ||
    document.querySelector(".dataset-letters") ||
    document.querySelector("[class*='dataset']");
  const correspondents = Array.from(document.querySelectorAll(".llc__item_correspondent, .ll-crpt"))
    .slice(0, 5)
    .map((el) => (el.innerText || el.textContent || "").trim())
    .filter(Boolean);
  const subjects = Array.from(document.querySelectorAll(".llc__subject, .ll-sj"))
    .slice(0, 5)
    .map((el) => (el.innerText || el.textContent || "").trim())
    .filter(Boolean);
  const resultCount = Math.max(correspondents.length, subjects.length);
  return {
    queryUrl: location.href,
    empty:
      resultCount === 0 &&
      (text.includes("Ничего не нашлось") || text.includes("Перейти во входящие")),
    unreadCount: datasetRoot
      ? datasetRoot.querySelectorAll(".llc__item_unread, .llc__subject_unread").length
      : 0,
    resultCount,
    bodyPreview: text.slice(0, 1200),
  };
}
"""
    )    


async def _probe_sent_listing(page, email: str) -> dict:
    await page.goto("https://e.mail.ru/sent/", wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1800)
    await page.evaluate(
        """
async (targetEmail) => {
  const lower = String(targetEmail || "").toLowerCase();
  for (let i = 0; i < 14; i += 1) {
    const text = (document.body.innerText || "").toLowerCase();
    if (text.includes(lower)) {
      return;
    }
    window.scrollTo(0, document.body.scrollHeight);
    await new Promise((resolve) => setTimeout(resolve, 450));
  }
}""",
        email,
    )


async def _probe_search_result_v2(page, query: str) -> dict:
    await page.goto("https://e.mail.ru/search/?q_query=" + urllib.parse.quote(query), wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1800)
    await page.evaluate(
        """
async (target) => {
  const lower = String(target || "").toLowerCase();
  for (let i = 0; i < 8; i += 1) {
    const text = (document.body.innerText || "").toLowerCase();
    if (text.includes(lower)) {
      return;
    }
    window.scrollTo(0, document.body.scrollHeight);
    await new Promise((resolve) => setTimeout(resolve, 400));
  }
}""",
        query,
    )
    return await page.evaluate(
        """
(targetEmail) => {
  const text = document.body.innerText || "";
  const lower = String(targetEmail || "").toLowerCase();
  const lines = text
    .split(/\\n+/)
    .map((item) => item.trim())
    .filter(Boolean);
  const datasetRoot =
    document.querySelector(".dataset__items") ||
    document.querySelector(".dataset-letters") ||
    document.querySelector("[class*='dataset']");
  const correspondents = Array.from(document.querySelectorAll(".llc__item_correspondent, .ll-crpt"))
    .slice(0, 5)
    .map((el) => (el.innerText || el.textContent || "").trim())
    .filter(Boolean);
  const subjects = Array.from(document.querySelectorAll(".llc__subject, .ll-sj"))
    .slice(0, 5)
    .map((el) => (el.innerText || el.textContent || "").trim())
    .filter(Boolean);
  const hitIndexes = [];
  for (let i = 0; i < lines.length; i += 1) {
    if (lines[i].toLowerCase().includes(lower)) {
      hitIndexes.push(i);
    }
  }
  const snippets = hitIndexes.map((index) => lines.slice(Math.max(0, index - 2), index + 8).join(" | "));
  const resultCount = Math.max(correspondents.length, subjects.length, hitIndexes.length);
  return {
    queryUrl: location.href,
    empty:
      hitIndexes.length === 0 &&
      (resultCount === 0 || text.includes("Ничего не нашлось") || text.includes("Перейти во входящие")),
    unreadCount: datasetRoot
      ? datasetRoot.querySelectorAll(".llc__item_unread, .llc__subject_unread").length
      : 0,
    resultCount,
    targetFound: hitIndexes.length > 0,
    hasIncoming: snippets.some((item) => item.toLowerCase().includes("входящие")),
    hasSent: snippets.some((item) => item.toLowerCase().includes("отправленные")),
    bodyPreview: snippets[0] || text.slice(0, 1200),
  };
}""",
        query,
    )
    return await page.evaluate(
        """(targetEmail) => {
  const text = document.body.innerText || "";
  const lower = String(targetEmail || "").toLowerCase();
  const items = text
    .split(/\\n+/)
    .map((item) => item.trim())
    .filter(Boolean);
  const hitIndex = items.findIndex((item) => item.toLowerCase().includes(lower));
  return {
    found: hitIndex >= 0 || text.toLowerCase().includes(lower),
    snippet: hitIndex >= 0 ? items.slice(Math.max(0, hitIndex - 1), hitIndex + 3).join(" | ") : "",
    pageUrl: location.href,
  };
}""",
        email,
    )


async def audit_mailru_inbox(project_root: Path, registry: list[ContactRecord], *, limit: int | None = None) -> list[AuditRecord]:
    state = MailOutreachState.load(_state_path(project_root))
    contacts = registry[:limit] if limit else registry
    client, playwright, browser, page, artifacts, logger = await _open_mail_session(project_root, label="mail_outreach_audit")
    await capture_screenshot(page, artifacts.screenshots_dir / "inbox.png", logger)
    audits: list[AuditRecord] = []
    try:
        for item in contacts:
            probe = await _probe_search_result_v2(page, item.email)
            sent_probe = _probe_sent_listing(page, item.email) if probe["empty"] else None
            sent_hit = await sent_probe if sent_probe else {"found": False, "snippet": "", "pageUrl": ""}
            local_sent = list(state.sent_messages.get(item.email, []))
            manual_known = "manual_already_contacted" in set(filter(None, item.source.split(",")))
            if probe.get("targetFound") and probe.get("hasIncoming"):
                if int(probe["unreadCount"] or 0) > 0:
                    status = "reply_unread"
                    detail = "Найден входящий тред от контакта, есть непрочитанный маркер."
                else:
                    status = "reply_read_confirmed"
                    detail = "Найден входящий тред от контакта без непрочитанного маркера."
            elif probe.get("targetFound") or sent_hit["found"] or local_sent or manual_known:
                status = "thread_found_sent_only"
                if probe.get("targetFound"):
                    detail = "Тред найден, но явного входящего письма от контакта не видно."
                elif sent_hit["found"]:
                    detail = "Глобальный поиск пустой, но адрес найден в списке Отправленные."
                elif local_sent:
                    detail = "Есть локальная запись об отправке, но Mail.ru не показал входящий тред."
                else:
                    detail = "Контакт отмечен как уже использованный вручную, но входящий тред не найден."
            else:
                status = "ready_to_send"
                detail = "Тред не найден, контакт выглядит как новый."
            audit = AuditRecord(
                contact_id=item.contact_id,
                email=item.email,
                brand_handle=item.brand_handle,
                brand_name=item.brand_name,
                instagram_url=item.instagram_url,
                status=status,
                detail=detail,
                query_url=str(probe["queryUrl"]),
                result_count=int(probe["resultCount"] or 0),
                unread_count=int(probe["unreadCount"] or 0),
                local_sent_count=len(local_sent),
                body_preview=str(probe["bodyPreview"] or sent_hit["snippet"] or ""),
                updated_at_iso=utcnow_iso(),
                preferred_contact_channel=item.preferred_contact_channel,
                is_ultra_premium_brand=item.is_ultra_premium_brand,
            )
            audits.append(audit)
            state.last_audit_by_email[item.email] = audit.to_dict()
            _write_brand_mail_summary(project_root, item, audit=audit)
    finally:
        state.last_audit_at_iso = utcnow_iso()
        state.save(_state_path(project_root))
        await _close_mail_session(client, playwright, browser)
    return audits


async def send_mailru_message(
    project_root: Path,
    *,
    to_email: str,
    subject: str,
    body: str,
    brand_handle: str = "",
    brand_name: str = "",
    instagram_url: str = "",
) -> dict:
    email = _sanitize_email(to_email)
    if not email:
        raise RuntimeError("Recipient email is required")
    client, playwright, browser, page, artifacts, logger = await _open_mail_session(project_root, label=f"mail_outreach_send_{_slug(email)}")
    before_shot = artifacts.screenshots_dir / "before_send.png"
    after_shot = artifacts.screenshots_dir / "after_send.png"
    try:
        await capture_screenshot(page, before_shot, logger)
        await page.goto("https://e.mail.ru/compose/", wait_until="domcontentloaded", timeout=60000)
        await page.wait_for_timeout(2000)

        recipient_input = page.locator("input.container--H9L5q").first
        await recipient_input.click(timeout=15000)
        await recipient_input.fill(email)
        await recipient_input.press("Enter")

        subject_input = page.locator("input[name='Subject']").first
        await subject_input.fill(subject)

        editor = page.locator("div[contenteditable='true'][role='textbox']").first
        await editor.click(timeout=10000)
        await editor.evaluate(
            """
(el, text) => {
  const lines = String(text || "").split(/\\n/);
  el.innerHTML = "";
  for (const line of lines) {
    const div = document.createElement("div");
    if (line) {
      div.textContent = line;
    } else {
      div.appendChild(document.createElement("br"));
    }
    el.appendChild(div);
  }
}
""",
            body,
        )

        await page.get_by_role("button", name=re.compile(r"^Отправить$", re.I)).first.click(timeout=15000)
        await page.wait_for_timeout(2500)
        await capture_screenshot(page, after_shot, logger)

        sent_payload = {
            "email": email,
            "brand_handle": brand_handle,
            "brand_name": brand_name,
            "instagram_url": instagram_url,
            "subject": subject,
            "body": body,
            "sent_at_iso": utcnow_iso(),
            "artifact_dir": str(artifacts.run_dir),
            "log_path": str(artifacts.log_path),
            "screenshot_before": str(before_shot),
            "screenshot_after": str(after_shot),
        }
        _append_jsonl(_mail_outreach_paths(project_root)["send_log"], sent_payload)
        state = MailOutreachState.load(_state_path(project_root))
        state.sent_messages.setdefault(email, []).append(sent_payload)
        state.last_send_at_iso = sent_payload["sent_at_iso"]
        state.save(_state_path(project_root))
        _append_manual_contacted_entry(
            project_root,
            email=email,
            instagram_url=instagram_url,
        )

        if brand_handle:
            contact = ContactRecord(
                contact_id=_slug(email),
                email=email,
                instagram_url=instagram_url,
                brand_handle=brand_handle,
                brand_name=brand_name,
                source="send_runtime",
                brand_folder=str(project_root / "output" / "brand_intelligence" / _slug(brand_handle)),
            )
            _write_brand_mail_summary(project_root, contact, sent_payload=sent_payload)
        return sent_payload
    finally:
        await _close_mail_session(client, playwright, browser)


async def run_mailru_cycle(
    project_root: Path,
    registry: list[ContactRecord],
    *,
    send_ready: bool,
    limit: int,
) -> dict:
    policy = load_farida_policy(project_root)
    state = MailOutreachState.load(_state_path(project_root))
    client, playwright, browser, page, artifacts, logger = await _open_mail_session(project_root, label="mail_outreach_cycle")
    try:
        async def audit_in_session() -> list[AuditRecord]:
            await page.goto("https://e.mail.ru/inbox/", wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(1200)
            await capture_screenshot(page, artifacts.screenshots_dir / "inbox.png", logger)
            audits_local: list[AuditRecord] = []
            for item in registry:
                probe = await _probe_search_result_v2(page, item.email)
                sent_probe = _probe_sent_listing(page, item.email) if probe["empty"] else None
                sent_hit = await sent_probe if sent_probe else {"found": False, "snippet": "", "pageUrl": ""}
                local_sent = list(state.sent_messages.get(item.email, []))
                manual_known = "manual_already_contacted" in set(filter(None, item.source.split(",")))
                if probe.get("targetFound") and probe.get("hasIncoming"):
                    if int(probe["unreadCount"] or 0) > 0:
                        status = "reply_unread"
                        detail = "Найден входящий тред от контакта, есть непрочитанный маркер."
                    else:
                        status = "reply_read_confirmed"
                        detail = "Найден входящий тред от контакта без непрочитанного маркера."
                elif probe.get("targetFound") or sent_hit["found"] or local_sent or manual_known:
                    status = "thread_found_sent_only"
                    if probe.get("targetFound"):
                        detail = "Тред найден, но явного входящего письма от контакта не видно."
                    elif sent_hit["found"]:
                        detail = "Глобальный поиск пустой, но адрес найден в списке Отправленные."
                    elif local_sent:
                        detail = "Есть локальная запись об отправке, но Mail.ru не показал входящий тред."
                    else:
                        detail = "Контакт отмечен как уже использованный вручную, но входящий тред не найден."
                else:
                    status = "ready_to_send"
                    detail = "Тред не найден, контакт выглядит как новый."
                audit = AuditRecord(
                    contact_id=item.contact_id,
                    email=item.email,
                    brand_handle=item.brand_handle,
                    brand_name=item.brand_name,
                    instagram_url=item.instagram_url,
                    status=status,
                    detail=detail,
                    query_url=str(probe["queryUrl"]),
                    result_count=int(probe["resultCount"] or 0),
                    unread_count=int(probe["unreadCount"] or 0),
                    local_sent_count=len(local_sent),
                    body_preview=str(probe["bodyPreview"] or sent_hit["snippet"] or ""),
                    updated_at_iso=utcnow_iso(),
                    preferred_contact_channel=item.preferred_contact_channel,
                    is_ultra_premium_brand=item.is_ultra_premium_brand,
                )
                audits_local.append(audit)
                state.last_audit_by_email[item.email] = audit.to_dict()
                _write_brand_mail_summary(project_root, item, audit=audit)
            state.last_audit_at_iso = utcnow_iso()
            state.save(_state_path(project_root))
            return audits_local

        audits = await audit_in_session()
        sent_count = 0
        if send_ready:
            audit_by_email = {item.email: item for item in audits}
            eligible = [
                item
                for item in registry
                if item.email and audit_by_email.get(item.email) and audit_by_email[item.email].status == "ready_to_send"
            ][:limit]
            for item in eligible:
                subject, body = build_master_message(
                    policy,
                    brand_name=item.brand_name or item.brand_handle or item.email,
                    channel="email",
                    seed_value=item.email,
                )
                email = _sanitize_email(item.email)
                before_shot = artifacts.screenshots_dir / f"before_send_{_slug(email)}.png"
                after_shot = artifacts.screenshots_dir / f"after_send_{_slug(email)}.png"

                await page.goto("https://e.mail.ru/compose/", wait_until="domcontentloaded", timeout=60000)
                await page.wait_for_timeout(2000)
                await capture_screenshot(page, before_shot, logger)

                recipient_input = page.locator("input.container--H9L5q").first
                await recipient_input.click(timeout=15000)
                await recipient_input.fill(email)
                await recipient_input.press("Enter")

                subject_input = page.locator("input[name='Subject']").first
                await subject_input.fill(subject)

                editor = page.locator("div[contenteditable='true'][role='textbox']").first
                await editor.click(timeout=10000)
                await editor.evaluate(
                    """
(el, text) => {
  const lines = String(text || "").split(/\\n/);
  el.innerHTML = "";
  for (const line of lines) {
    const div = document.createElement("div");
    if (line) {
      div.textContent = line;
    } else {
      div.appendChild(document.createElement("br"));
    }
    el.appendChild(div);
  }
}
""",
                    body,
                )

                await page.get_by_role("button", name=re.compile(r"^Отправить$", re.I)).first.click(timeout=15000)
                await page.wait_for_timeout(2500)
                await capture_screenshot(page, after_shot, logger)

                sent_payload = {
                    "email": email,
                    "brand_handle": item.brand_handle,
                    "brand_name": item.brand_name,
                    "instagram_url": item.instagram_url,
                    "subject": subject,
                    "body": body,
                    "sent_at_iso": utcnow_iso(),
                    "artifact_dir": str(artifacts.run_dir),
                    "log_path": str(artifacts.log_path),
                    "screenshot_before": str(before_shot),
                    "screenshot_after": str(after_shot),
                }
                _append_jsonl(_mail_outreach_paths(project_root)["send_log"], sent_payload)
                state.sent_messages.setdefault(email, []).append(sent_payload)
                state.last_send_at_iso = sent_payload["sent_at_iso"]
                state.save(_state_path(project_root))
                _append_manual_contacted_entry(
                    project_root,
                    email=email,
                    instagram_url=item.instagram_url,
                )
                _write_brand_mail_summary(project_root, item, sent_payload=sent_payload)
                sent_count += 1

            audits = await audit_in_session()

        return {"audits": audits, "sent_count": sent_count, "artifact_dir": str(artifacts.run_dir)}
    finally:
        await _close_mail_session(client, playwright, browser)
