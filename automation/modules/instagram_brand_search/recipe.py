from __future__ import annotations

from dataclasses import asdict
from datetime import datetime, timedelta, timezone
from html import unescape
from pathlib import Path
import json
import logging
import os
import re
import shutil
import zipfile
from urllib.parse import urlparse
from xml.sax.saxutils import escape as xml_escape

from playwright.async_api import BrowserContext, Locator, Page, TimeoutError as PlaywrightTimeoutError

from ...browser import capture_screenshot
from ...human import HumanSettings, Humanizer
from .models import BloggerRunStats, BloggerTarget, BrandAssessment, FollowingCandidate, MentionCandidate, PostSnapshot
from .state import InstagramBrandSearchState


INSTAGRAM_HOME_URL = "https://www.instagram.com/"
INSTAGRAM_WEB_APP_ID = "936619743392459"
POST_LINK_RE = re.compile(r"instagram\.com/(?:[A-Za-z0-9._]+/)?(?:p|reel)/([A-Za-z0-9_-]+)/?", re.IGNORECASE)
PROFILE_LINK_RE = re.compile(r"instagram\.com/([A-Za-z0-9._]+)/?$", re.IGNORECASE)
HANDLE_RE = re.compile(r"(?<![\w.])@([A-Za-z0-9._]{1,30})")
AD_KEYWORDS = (
    "реклама",
    "промокод",
    "скидк",
    "заказ",
    "shop",
    "store",
    "collection",
    "brand",
    "collab",
    "partnership",
    "sponsored",
    "partner",
    "thanks to",
    "wearing",
    "look from",
    "dress by",
    "shop now",
)
BRAND_KEYWORDS = (
    "official",
    "shop",
    "store",
    "atelier",
    "collection",
    "denim",
    "beauty",
    "cosmetic",
    "cosmetics",
    "boutique",
    "supplement",
    "nutrition",
    "fit",
    "sport",
    "flower",
    "flowers",
    "studio",
    "brand",
    "wear",
    "app",
    "apps",
    "marketplace",
    "delivery",
    "taxi",
    "fintech",
    "jewelry",
    "jewellery",
    "hotel",
    "travel",
    "decor",
    "home",
    "fmcg",
    "бренд",
    "магазин",
    "салон",
    "клиника",
    "ресторан",
    "кафе",
    "доставка",
    "массаж",
    "косметолог",
    "студия",
)
PERSON_KEYWORDS = (
    "mom",
    "blogger",
    "personal",
    "photographer",
    "traveler",
    "life",
    "journal",
    "model",
    "influencer",
    "creator",
    "stylist",
    "trainer",
    "coach",
    "photography",
    "фотограф",
    "блогер",
    "модель",
    "тренер",
    "коуч",
    "стилист",
    "психолог",
    "врач",
    "доктор",
    "ортопед",
    "хирург",
)
SERVICE_KEYWORDS = (
    "photographer",
    "photography",
    "beauty",
    "makeup",
    "stylist",
    "hair",
    "trainer",
    "coach",
    "doctor",
    "clinic",
    "massage",
    "cosmetolog",
    "delivery",
    "taxi",
    "booking",
    "app",
    "marketplace",
    "fintech",
    "hotel",
    "travel",
    "restaurant",
    "cafe",
    "decor",
    "interior",
    "косметолог",
    "визажист",
    "макияж",
    "фотограф",
    "стилист",
    "парикмахер",
    "тренер",
    "коуч",
    "врач",
    "доктор",
    "клиника",
    "массаж",
    "съёмки",
    "съемки",
    "брендов",
)
PROJECT_KEYWORDS = (
    "charity",
    "shelter",
    "media",
    "project",
    "animal",
    "rescue",
    "благотвор",
    "приют",
    "проект",
    "медиа",
    "спасение",
)


FEMALE_PROFILE_KEYWORDS = (
    "she/her",
    "girl",
    "woman",
    "women",
    "mom",
    "mama",
    "mother",
    "wife",
    "actress",
    "model",
    "beauty",
    "makeup",
    "fashion",
    "блогер",
    "мама",
    "девушка",
    "женщина",
    "актриса",
    "модель",
    "стилист",
    "визажист",
    "косметолог",
)
MALE_PROFILE_KEYWORDS = (
    "he/him",
    "dad",
    "father",
    "husband",
    "man",
    "men",
    "mr",
    "coach",
    "предприниматель",
    "муж",
    "отец",
    "мужчина",
    "парень",
    "актер",
    "ведущий",
)
COMMON_FEMALE_NAMES = {
    "alina", "alisa", "amina", "anastasia", "angelina", "anna", "arina", "bella", "daria", "diana",
    "ekaterina", "elena", "eva", "karina", "katya", "kira", "liza", "maria", "marina", "milana",
    "natalia", "olga", "polina", "sabina", "sofia", "sonya", "sveta", "tatiana", "uliana", "vera",
    "vika", "victoria", "yana", "yulia", "zarina", "алена", "алина", "алиса", "анастасия", "ангелина",
    "анна", "арина", "валерия", "вера", "вика", "виктория", "дарья", "диана", "екатерина", "елена",
    "ирина", "карина", "катя", "кристина", "ксения", "лиза", "марина", "мария", "милана", "наталья",
    "оксана", "ольга", "полина", "сабина", "света", "светлана", "софия", "софья", "татьяна", "ульяна",
    "фатима", "элина", "юлия", "яна",
}


PRIORITY_NICHE_KEYWORDS: dict[str, tuple[str, ...]] = {
    "beauty": (
        "beauty", "skincare", "skin care", "makeup", "cosmetic", "cosmetics", "haircare", "hair care",
        "bodycare", "body care", "perfume", "fragrance", "parfum", "beauty gadget", "уход", "космет",
        "макияж", "парфюм", "аромат", "волос", "волосы", "кожа", "уход за кожей",
    ),
    "fashion": (
        "fashion", "style", "wear", "outfit", "wardrobe", "dress", "lingerie", "underwear", "loungewear",
        "shoes", "bag", "bags", "accessories", "jewelry", "jewellery", "homewear", "clothes", "clothing",
        "обув", "одеж", "сумк", "аксессуар", "украш", "бель", "домашн", "стиль", "лук",
    ),
    "apps_services": (
        "app", "apps", "mobile app", "service", "services", "subscription", "delivery", "taxi", "marketplace",
        "booking", "editor", "editing", "photo app", "video app", "fintech", "wallet", "banking",
        "маркетплейс", "доставка", "такси", "сервис", "прилож", "подписк", "запись", "бронь", "монтаж", "редактор",
    ),
    "fmcg_everyday": (
        "drink", "beverage", "snack", "food", "groceries", "household", "hygiene", "soap", "detergent",
        "toothpaste", "everyday", "fmcg", "напит", "снек", "продукт", "гигиен", "бытов", "ежедневн",
    ),
    "sports_wellness": (
        "fitness", "wellness", "activewear", "sportswear", "workout", "gym", "healthy food", "protein",
        "supplement", "nutrition", "recovery", "training", "sport", "фитнес", "трениров", "спорт",
        "спортив", "велнес", "нутри", "здоров",
    ),
    "jewelry_accessories": (
        "jewelry", "jewellery", "accessories", "watch", "watches", "rings", "earrings", "bracelet",
        "necklace", "украш", "серьг", "кольц", "браслет", "аксессуар", "часы",
    ),
    "clinics_cosmetology_dentistry": (
        "clinic", "cosmetology", "cosmetologist", "dental", "dentist", "aesthetic", "injector",
        "клиник", "косметолог", "стомат", "дантист", "эстет", "инъекц",
    ),
    "travel_leisure": (
        "travel", "hotel", "hotels", "resort", "restaurant", "restaurants", "cafe", "leisure", "vacation",
        "spa", "booking", "trip", "путешеств", "отел", "ресторан", "кафе", "курорт", "отдых",
    ),
    "home_decor": (
        "home", "decor", "interior", "household", "kitchen", "home goods", "furniture", "candles", "bedding",
        "декор", "интерьер", "дом", "кухня", "товары для дома", "мебель", "посуда", "текстиль",
    ),
}
FOLLOWING_BRAND_REJECT_KEYWORDS = (
    "official", "shop", "store", "boutique", "atelier", "showroom", "brand", "marketplace", "delivery", "taxi",
    "app", "mobile app", "booking", "book now", "service", "customer support", "support", "clinic", "dental",
    "dentist", "salon", "restaurant", "cafe", "hotel", "resort", "menu", "open daily", "open every day",
    "order", "orders", "заказ", "заказы", "доставка", "салон", "клиника", "стомат", "ресторан", "кафе",
    "отель", "меню", "запись", "записаться", "market",
)
FOLLOWING_BRAND_HANDLE_HINTS = (
    "shop", "store", "official", "brand", "boutique", "atelier", "salon", "clinic", "dental", "hotel",
    "cafe", "restaurant", "delivery", "taxi", "app", "market", "marketplace",
)
FOLLOWING_CREATOR_KEYWORDS = (
    "blogger", "creator", "content creator", "ugc", "lifestyle", "personal blog", "reels", "humor", "family",
    "mom", "mother", "model", "actress", "artist", "singer", "influencer", "fitness", "wellness",
    "блогер", "креатор", "лайфстайл", "личный блог", "рилс", "юмор", "семья", "мама", "модель", "актриса",
    "певица", "инфлюенсер",
)


def _compact(text: str) -> str:
    return " ".join((text or "").split())


def _lower_set(items: list[str]) -> set[str]:
    return {item.lower() for item in items if item}


def normalize_text(value: str) -> str:
    text = unescape((value or "").replace("\r\n", "\n").replace("\r", "\n"))
    text = text.replace("\u00a0", " ").replace("\u200b", "")
    return text.strip()


def md_link(label: str, target: str) -> str:
    safe_label = label.replace("[", "\\[").replace("]", "\\]")
    safe_target = (target or "").replace(" ", "%20")
    return f"[{safe_label}]({safe_target})"


def relative_markdown_target(report_path: Path, target: str) -> str:
    if not target:
        return ""
    target_path = Path(target)
    if not target_path.is_absolute():
        return target_path.as_posix()
    try:
        rel = os.path.relpath(target_path, report_path.parent)
        return Path(rel).as_posix()
    except ValueError:
        return target_path.as_posix()


def canonical_handle(handle: str) -> str:
    return re.sub(r"[._]+", "", (handle or "").strip().lower())


EXPORT_BUSINESS_MARKERS = (
    "brand",
    "store",
    "shop",
    "official",
    "boutique",
    "collection",
    "atelier",
    "studio",
    "showroom",
    "salon",
    "clinic",
    "app",
    "apps",
    "marketplace",
    "concierge",
    "hotel",
    "restaurant",
    "cafe",
    "club",
    "agency",
    "jewelry",
    "jewellery",
    "beauty",
    "cosmetic",
    "cosmetics",
    "wear",
    "lingerie",
    "swimwear",
    "furniture",
    "decor",
    "flowers",
    "fashion",
    "мaгaз",
    "магаз",
    "бренд",
    "салон",
    "студ",
    "клиник",
    "ателье",
    "доставка",
    "прилож",
)

EXPORT_HARD_BUSINESS_MARKERS = (
    "store",
    "shop",
    "official",
    "boutique",
    "collection",
    "atelier",
    "studio",
    "showroom",
    "salon",
    "clinic",
    "app",
    "apps",
    "marketplace",
    "concierge",
    "hotel",
    "restaurant",
    "cafe",
    "agency",
    "wear",
    "lingerie",
    "swimwear",
    "flowers",
    "магаз",
    "салон",
    "клиник",
    "ателье",
    "прилож",
)

EXPORT_PERSON_ROLE_MARKERS = (
    "pr",
    "marketing",
    "influence marketing",
    "producer",
    "stylist",
    "hair",
    "mua",
    "makeup",
    "photographer",
    "videographer",
    "videomaker",
    "retoucher",
    "artist",
    "model",
    "blogger",
    "creator",
    "актриса",
    "модель",
    "стилист",
    "фотограф",
    "видеограф",
    "ретуш",
    "продюсер",
    "маркетинг",
    "маркетолог",
    "блогер",
)


def has_any_marker(text: str, markers: tuple[str, ...]) -> bool:
    lowered = normalize_text(text).lower()
    if not lowered:
        return False
    return any(marker in lowered for marker in markers)


def looks_like_person_display_name(display_name: str) -> bool:
    normalized = normalize_text(display_name)
    if not normalized:
        return False
    candidate = normalized.split("|", 1)[0].strip()
    if has_any_marker(candidate, EXPORT_BUSINESS_MARKERS):
        return False
    parts = [part for part in re.split(r"[\s/]+", candidate) if part]
    if len(parts) < 2 or len(parts) > 4:
        return False
    if any(re.search(r"\d", part) for part in parts):
        return False
    word_like = [part for part in parts if re.fullmatch(r"[A-Za-zА-Яа-яЁё'_-]{2,}", part)]
    return len(word_like) == len(parts)


def is_probable_person_brand_false_positive(record: dict) -> bool:
    handle = str(record.get("handle", ""))
    display_name = str(record.get("display_name", ""))
    bio = str(record.get("bio", ""))
    category_label = str(record.get("category_label", ""))
    external_link = str(record.get("external_link", ""))
    account_kind = str(record.get("account_kind", ""))
    reasoning = str(record.get("reasoning", "")) or str(record.get("brand_reasoning", ""))

    strong_business_signal = any(
        has_any_marker(value, EXPORT_HARD_BUSINESS_MARKERS)
        for value in (handle, display_name, external_link)
    )
    person_role_signal = any(
        has_any_marker(value, EXPORT_PERSON_ROLE_MARKERS)
        for value in (display_name, bio, category_label, reasoning)
    )
    name_signal = looks_like_person_display_name(display_name)
    handle_signal = bool(re.fullmatch(r"[a-zа-я0-9]+(?:[._][a-zа-я0-9]+)+", (handle or "").lower()))

    if account_kind == "service_provider":
        return True
    if name_signal and not strong_business_signal:
        return True
    if person_role_signal and not strong_business_signal:
        return True
    if handle_signal and person_role_signal and not strong_business_signal:
        return True
    return False


def is_exportable_brand_record(record: dict) -> bool:
    if not record:
        return False
    if not (record.get("is_brand") or record.get("is_brand_like")):
        return False
    if is_probable_person_brand_false_positive(record):
        return False
    account_kind = str(record.get("account_kind", ""))
    outreach_fit = str(record.get("outreach_fit", "low"))
    return outreach_fit != "low" or account_kind in {"brand_store", "service_provider"} or bool(record.get("is_brand_like"))


def dedupe_records_by_handle(records: list[dict]) -> list[dict]:
    deduped: dict[str, dict] = {}
    for record in records:
        key = canonical_handle(str(record.get("handle", "")))
        if not key:
            continue
        current = deduped.get(key)
        if current is None:
            deduped[key] = record
            continue
        current_score = (
            int(current.get("followers_count", 0) or 0),
            len(current.get("sources", []) or []),
        )
        incoming_score = (
            int(record.get("followers_count", 0) or 0),
            len(record.get("sources", []) or []),
        )
        if incoming_score > current_score:
            deduped[key] = record
    return list(deduped.values())


def dedupe_brand_sources(sources: list[dict]) -> list[dict]:
    deduped: dict[tuple[str, str, str], dict] = {}
    for source in sources or []:
        blogger_handle = canonical_handle(str(source.get("blogger_handle", "")))
        post_url = str(source.get("post_url", "")).strip()
        post_date_iso = str(source.get("post_date_iso", "")).strip()
        key = (blogger_handle, post_url, post_date_iso)
        current = deduped.get(key)
        if current is None:
            deduped[key] = dict(source)
            continue

        for field in ("ad_likelihood", "ad_reasoning", "caption_excerpt"):
            current_value = str(current.get(field, "") or "")
            incoming_value = str(source.get(field, "") or "")
            if len(incoming_value) > len(current_value):
                current[field] = incoming_value

    return sorted(
        deduped.values(),
        key=lambda item: (
            str(item.get("post_date_iso", "") or ""),
            str(item.get("blogger_handle", "") or ""),
            str(item.get("post_url", "") or ""),
        ),
    )


def collect_exportable_brand_records(state: InstagramBrandSearchState) -> list[tuple[str, dict]]:
    merged: dict[str, dict] = {}
    for record in state.brand_records.values():
        if not is_exportable_brand_record(record):
            continue
        canonical = canonical_handle(str(record.get("handle", "")))
        if not canonical:
            continue
        display_handle = str(record.get("handle", "")).strip().lstrip("@")
        if not display_handle:
            display_handle = extract_handle_from_url(str(record.get("profile_url", ""))) or canonical

        current = merged.get(canonical)
        if current is None:
            item = dict(record)
            item["handle"] = display_handle
            item["_canonical_handle"] = canonical
            item["sources"] = dedupe_brand_sources(record.get("sources", []))
            merged[canonical] = item
            continue

        current_sources = current.get("sources", []) or []
        incoming_sources = record.get("sources", []) or []
        current["sources"] = dedupe_brand_sources(current_sources + incoming_sources)
        if not str(current.get("handle", "") or "").strip() and display_handle:
            current["handle"] = display_handle

        for field in (
            "profile_url",
            "display_name",
            "bio",
            "category_label",
            "external_link",
            "niche",
            "reasoning",
            "account_kind",
            "outreach_fit",
            "brand_likelihood",
            "ad_likelihood",
            "screenshot_path",
        ):
            current_value = str(current.get(field, "") or "")
            incoming_value = str(record.get(field, "") or "")
            if not current_value and incoming_value:
                current[field] = incoming_value

        if int(record.get("followers_count", 0) or 0) > int(current.get("followers_count", 0) or 0):
            current["followers_count"] = int(record.get("followers_count", 0) or 0)
            if record.get("followers_text"):
                current["followers_text"] = record.get("followers_text", "")

        current["is_brand"] = bool(current.get("is_brand")) or bool(record.get("is_brand"))
        current["is_brand_like"] = bool(current.get("is_brand_like")) or bool(record.get("is_brand_like"))

    return sorted(
        [
            (str(record.get("handle", "")).strip().lstrip("@") or canonical, record)
            for canonical, record in merged.items()
        ],
        key=lambda item: item[0].lower(),
    )


def merge_following_records_by_handle(records: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for record in records:
        key = canonical_handle(str(record.get("handle", "")))
        if not key:
            continue
        current = merged.get(key)
        if current is None:
            item = dict(record)
            source_handle = str(record.get("source_blogger_handle", "")).strip()
            source_url = str(record.get("source_blogger_url", "")).strip()
            item["_source_handles"] = {source_handle} if source_handle else set()
            item["_source_urls"] = {source_url} if source_url else set()
            item["_occurrences"] = 1
            merged[key] = item
            continue

        source_handle = str(record.get("source_blogger_handle", "")).strip()
        source_url = str(record.get("source_blogger_url", "")).strip()
        if source_handle:
            current["_source_handles"].add(source_handle)
        if source_url:
            current["_source_urls"].add(source_url)
        current["_occurrences"] += 1

        current_score = (
            int(current.get("is_selected_target", False)),
            int(current.get("followers_count", 0) or 0),
            int(current.get("is_brand_like", False)),
        )
        incoming_score = (
            int(record.get("is_selected_target", False)),
            int(record.get("followers_count", 0) or 0),
            int(record.get("is_brand_like", False)),
        )
        if incoming_score > current_score:
            carry_handles = current["_source_handles"]
            carry_urls = current["_source_urls"]
            carry_occurrences = current["_occurrences"]
            current.clear()
            current.update(record)
            current["_source_handles"] = carry_handles
            current["_source_urls"] = carry_urls
            current["_occurrences"] = carry_occurrences

    result = list(merged.values())
    for item in result:
        item["_source_handles"] = sorted(item.get("_source_handles", set()))
        item["_source_urls"] = sorted(item.get("_source_urls", set()))
    return result


def get_following_selected_follower_bounds(job: dict) -> tuple[int, int]:
    policy = job.get("following_scan_policy", {})
    min_followers = int(policy.get("follower_threshold", 300000) or 0)
    max_followers = int(policy.get("max_selected_followers", 0) or 0)
    return min_followers, max_followers


def followers_within_selected_range(followers_count: int, job: dict) -> bool:
    min_followers, max_followers = get_following_selected_follower_bounds(job)
    if followers_count < min_followers:
        return False
    if max_followers and followers_count > max_followers:
        return False
    return True


def is_selected_following_target_record(record: dict, job: dict) -> bool:
    followers_count = int(record.get("followers_count", 0) or 0)
    reject_brand_like = bool(job.get("following_scan_policy", {}).get("reject_brand_like_profiles", True))
    is_female_candidate = bool(record.get("is_female_candidate"))
    is_brand_like = bool(record.get("is_brand_like"))
    return (
        followers_within_selected_range(followers_count, job)
        and is_female_candidate
        and (not reject_brand_like or not is_brand_like)
    )


def build_phase1_shortlist_paths(job: dict) -> tuple[Path, Path]:
    base_dir = Path(job["outputs"]["following_candidates_dir"])
    return (
        base_dir / "shortlisted_bloggers_for_phase1.md",
        base_dir / "shortlisted_bloggers_for_phase1.txt",
    )


def normalize_instagram_url(raw_url: str) -> str:
    text = raw_url.strip()
    if not text:
        return ""
    if not text.startswith("http"):
        text = f"https://www.instagram.com/{text.lstrip('@').strip('/')}/"
    parsed = urlparse(text)
    cleaned = f"https://www.instagram.com{parsed.path}"
    if not cleaned.endswith("/"):
        cleaned += "/"
    return cleaned


def extract_shortcode(post_url: str) -> str:
    match = POST_LINK_RE.search(post_url or "")
    return match.group(1) if match else ""


def extract_handle_from_url(profile_url: str) -> str:
    match = PROFILE_LINK_RE.search((profile_url or "").rstrip("/"))
    return match.group(1) if match else ""


def target_queue_key(target: BloggerTarget) -> str:
    return normalize_instagram_url(target.profile_url)


def following_candidate_key(source_blogger_handle: str, candidate_handle: str) -> str:
    return f"{canonical_handle(source_blogger_handle)}::{canonical_handle(candidate_handle)}"


def safe_handle_slug(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (value or "").strip())
    return cleaned.strip("._-") or "profile"


def parse_compact_number(raw_text: str) -> int:
    text = normalize_text(raw_text).lower().replace("\u202f", " ").replace("\xa0", " ")
    text = text.replace("followers", "").replace("following", "").replace("posts", "")
    text = text.replace("подписчиков", "").replace("подписки", "").replace("подписок", "").replace("публикаций", "")
    text = text.strip()
    if not text:
        return 0

    suffix = ""
    suffix_match = re.search(r"(k|m|b|тыс\.?|млн|млрд)$", text)
    if suffix_match:
        suffix = suffix_match.group(1)
        text = text[: suffix_match.start()].strip()

    normalized_numeric = text.replace(" ", "")
    compact_numeric = re.fullmatch(r"\d+(?:[.,]\d+)?", normalized_numeric)
    if compact_numeric:
        if suffix:
            base = float(normalized_numeric.replace(",", "."))
        elif "," in normalized_numeric and "." not in normalized_numeric and len(normalized_numeric.rsplit(",", 1)[-1]) == 3:
            base = float(normalized_numeric.replace(",", ""))
        elif "." in normalized_numeric and "," not in normalized_numeric and len(normalized_numeric.rsplit(".", 1)[-1]) == 3:
            base = float(normalized_numeric.replace(".", ""))
        else:
            base = float(normalized_numeric.replace(",", "."))
    else:
        digits_only = re.sub(r"[^\d]", "", text)
        if not digits_only:
            return 0
        if suffix:
            base = float(digits_only)
        else:
            return int(digits_only)

    multiplier = 1
    if suffix == "k" or suffix.startswith("тыс"):
        multiplier = 1_000
    elif suffix == "m" or suffix == "млн":
        multiplier = 1_000_000
    elif suffix == "b" or suffix == "млрд":
        multiplier = 1_000_000_000
    return int(base * multiplier)


def extract_followers_count(*texts: str) -> int:
    patterns = (
        re.compile(r"([\d][\d\s.,]*\s*(?:k|m|b|тыс\.?|млн|млрд)?)\s*(?:followers?|подписчик[а-я]*)", re.IGNORECASE),
        re.compile(r"(?:followers?|подписчик[а-я]*)\s*[:\-]?\s*([\d][\d\s.,]*\s*(?:k|m|b|тыс\.?|млн|млрд)?)", re.IGNORECASE),
    )
    for text in texts:
        normalized = normalize_text(text)
        if not normalized:
            continue
        for pattern in patterns:
            match = pattern.search(normalized)
            if match:
                count = parse_compact_number(match.group(1))
                if count:
                    return count
    return 0


def tokenize_profile_identity(*parts: str) -> list[str]:
    haystack = normalize_text(" ".join(part for part in parts if part)).lower()
    return [token for token in re.split(r"[^a-zа-я0-9]+", haystack, flags=re.IGNORECASE) if token]


def classify_female_profile(*, handle: str, display_name: str, bio: str, category_label: str) -> tuple[bool, str, str]:
    haystack = normalize_text(" ".join([handle, display_name, bio, category_label])).lower()
    score = 0
    reasons: list[str] = []

    female_hits = [keyword for keyword in FEMALE_PROFILE_KEYWORDS if keyword in haystack]
    male_hits = [keyword for keyword in MALE_PROFILE_KEYWORDS if keyword in haystack]
    if female_hits:
        score += 1
        reasons.append(f"female-keywords: {', '.join(female_hits[:3])}")
    if male_hits:
        score -= 2
        reasons.append(f"male-keywords: {', '.join(male_hits[:3])}")

    tokens = tokenize_profile_identity(display_name, handle)
    matched_names = [token for token in tokens if token in COMMON_FEMALE_NAMES]
    if matched_names:
        score += 2
        reasons.append(f"female-name: {matched_names[0]}")

    if re.search(r"\b(she|her)\b", haystack):
        score += 2
    if re.search(r"\b(he|him)\b", haystack):
        score -= 2

    if score >= 3:
        return True, "high", "; ".join(reasons[:4]) or "female-coded profile signals"
    if score >= 2:
        return True, "medium", "; ".join(reasons[:4]) or "some female profile signals"
    return False, "low", "; ".join(reasons[:4]) or "not enough female-profile signals"


def detect_priority_niche(*parts: str) -> str:
    haystack = normalize_text(" ".join(part for part in parts if part)).lower()
    best_niche = ""
    best_score = 0
    for niche, keywords in PRIORITY_NICHE_KEYWORDS.items():
        score = sum(1 for keyword in keywords if keyword in haystack)
        if score > best_score:
            best_niche = niche
            best_score = score
    return best_niche


def classify_following_brand_exclusion(
    *,
    handle: str,
    display_name: str,
    bio: str,
    category_label: str,
    external_link: str,
) -> tuple[bool, str, str, str]:
    haystack = normalize_text(" ".join([handle, display_name, bio, category_label])).lower()
    brand_score = 0
    creator_score = 0
    reasons: list[str] = []

    matched_brand_keywords = [keyword for keyword in FOLLOWING_BRAND_REJECT_KEYWORDS if keyword in haystack]
    if matched_brand_keywords:
        brand_score += min(3, len(matched_brand_keywords))
        reasons.append(f"business-keywords: {', '.join(matched_brand_keywords[:3])}")

    handle_lower = handle.lower()
    handle_hits = [keyword for keyword in FOLLOWING_BRAND_HANDLE_HINTS if keyword in handle_lower]
    if handle_hits:
        brand_score += 2
        reasons.append(f"brand-handle: {', '.join(handle_hits[:2])}")

    if external_link:
        brand_score += 1
        reasons.append("external-link-present")

    creator_hits = [keyword for keyword in FOLLOWING_CREATOR_KEYWORDS if keyword in haystack]
    if creator_hits:
        creator_score += min(3, len(creator_hits))
        reasons.append(f"creator-keywords: {', '.join(creator_hits[:3])}")

    female_hits = [keyword for keyword in FEMALE_PROFILE_KEYWORDS if keyword in haystack]
    if female_hits:
        creator_score += 1
    person_hits = [keyword for keyword in PERSON_KEYWORDS if keyword in haystack]
    if person_hits:
        creator_score += 1

    tokens = tokenize_profile_identity(display_name, handle)
    if any(token in COMMON_FEMALE_NAMES for token in tokens):
        creator_score += 2

    niche = detect_priority_niche(display_name, bio, category_label, handle)

    if brand_score >= 4:
        return True, "high", "; ".join(reasons[:5]) or "strong business profile signals", niche
    if brand_score >= 2 and creator_score <= 1:
        return True, "medium", "; ".join(reasons[:5]) or "more business than creator signals", niche
    return False, "low", "; ".join(reasons[:5]) or "no strong business profile signals", niche


def candidate_key(post_url: str, handle: str) -> str:
    return f"{post_url}::{handle.lower()}"


def should_skip_line(raw_line: str) -> bool:
    stripped = raw_line.strip()
    return not stripped or stripped.startswith("#")


def seed_target_limit(job: dict) -> int:
    return int(job.get("inputs", {}).get("max_seed_targets", 0) or 0)


def classify_seed_source_kind(handle: str) -> str:
    lowered = canonical_handle(handle)
    brand_hints = (
        "shop",
        "store",
        "brand",
        "boutique",
        "atelier",
        "official",
        "collection",
        "limited",
        "showroom",
        "kids",
    )
    return "seed_brand" if any(hint in lowered for hint in brand_hints) else "seed"


def load_blogger_targets(path: Path, *, limit: int = 0) -> list[BloggerTarget]:
    targets: list[BloggerTarget] = []
    seen_urls: set[str] = set()
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        if should_skip_line(raw_line):
            continue
        url = normalize_instagram_url(raw_line)
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)
        handle = extract_handle_from_url(url)
        targets.append(
            BloggerTarget(
                profile_url=url,
                handle=handle,
                source_kind=classify_seed_source_kind(handle),
            )
        )
        if limit and len(targets) >= limit:
            break
    return targets


def build_humanizer(page: Page, job: dict) -> Humanizer:
    policy = job.get("humanization_policy", {})
    settings = HumanSettings(
        pause_range_ms=(policy.get("min_pause_ms", 70), policy.get("max_pause_ms", 450)),
        think_range_ms=(90, 220),
        click_delay_ms=(18, 45),
        type_delay_ms=(12, 28),
        move_steps=(6, 11),
        click_jitter_ratio=0.08,
        stability_checks=2,
        stability_interval_ms=90,
    )
    return Humanizer(page, settings=settings)


async def dismiss_instagram_popups(page: Page, human: Humanizer, logger: logging.Logger) -> None:
    patterns = [
        r"^Not now$",
        r"^Not Now$",
        r"^Not now\.$",
        r"^Не сейчас$",
        r"^Сейчас не$",
        r"^Close$",
        r"^Cancel$",
        r"^Dismiss$",
    ]
    for pattern in patterns:
        button = page.locator("button").filter(has_text=re.compile(pattern, re.IGNORECASE)).first
        if await button.count():
            try:
                if await button.is_visible():
                    logger.info("Closing Instagram popup with button pattern %s", pattern)
                    await human.human_click(button)
                    await human.pause(350, 800)
            except Exception:
                logger.info("Popup disappeared while closing")

    try:
        notification_dialog = page.locator('[role="dialog"]').filter(
            has_text=re.compile(r"turn on notifications|notifications|уведом", re.IGNORECASE)
        ).last
        if await notification_dialog.count() and await notification_dialog.is_visible():
            dismiss_button = notification_dialog.locator("button").filter(
                has_text=re.compile(r"not now|не сейчас|сейчас не|close|cancel|dismiss", re.IGNORECASE)
            ).first
            if await dismiss_button.count() and await dismiss_button.is_visible():
                logger.info("Closing Instagram notification prompt")
                await human.human_click(dismiss_button)
                await human.pause(450, 900)
    except Exception:
        logger.info("Notification popup check failed; continuing")


async def profile_page_has_posts(page: Page) -> bool:
    try:
        count = await page.locator('a[href*="/p/"], a[href*="/reel/"]').count()
        return count > 0
    except Exception:
        return False


async def ensure_profile_page(page: Page, human: Humanizer, logger: logging.Logger, profile_url: str) -> None:
    current = normalize_instagram_url(page.url) if page.url else ""
    if current == profile_url and await profile_page_has_posts(page):
        await dismiss_instagram_popups(page, human, logger)
        return

    await page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
    await human.pause(1500, 2400)
    await dismiss_instagram_popups(page, human, logger)
    for _ in range(5):
        if await profile_page_has_posts(page):
            return
        await human.pause(800, 1200)
    logger.info("Profile grid not ready after initial navigation; reloading %s", profile_url)
    await page.reload(wait_until="domcontentloaded", timeout=60000)
    await human.pause(1500, 2200)
    await dismiss_instagram_popups(page, human, logger)


async def ensure_profile_loaded(page: Page, human: Humanizer, logger: logging.Logger, profile_url: str) -> None:
    current = normalize_instagram_url(page.url) if page.url else ""
    if current == profile_url:
        await dismiss_instagram_popups(page, human, logger)
        return

    await page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
    await human.pause(1200, 2000)
    await dismiss_instagram_popups(page, human, logger)


async def return_to_profile_from_post(page: Page, human: Humanizer, logger: logging.Logger, profile_url: str) -> None:
    current = normalize_instagram_url(page.url) if page.url else ""
    if current == profile_url and await profile_page_has_posts(page):
        return

    for _ in range(2):
        try:
            await page.go_back(wait_until="domcontentloaded", timeout=30000)
            await human.pause(700, 1200)
            if normalize_instagram_url(page.url) == profile_url and await profile_page_has_posts(page):
                await dismiss_instagram_popups(page, human, logger)
                return
        except Exception:
            break

    await ensure_profile_page(page, human, logger, profile_url)


async def visible_post_links(page: Page) -> list[tuple[str, bool]]:
    anchors = page.locator('a[href*="/p/"], a[href*="/reel/"]')
    count = await anchors.count()
    results: list[tuple[str, bool]] = []
    seen: set[str] = set()
    for index in range(min(count, 24)):
        anchor = anchors.nth(index)
        try:
            if not await anchor.is_visible():
                continue
            href = await anchor.get_attribute("href")
            if not href:
                continue
            if href.startswith("/"):
                href = f"https://www.instagram.com{href}"
            href = href.split("?", 1)[0].rstrip("/") + "/"
            if href in seen:
                continue
            seen.add(href)
            html = (await anchor.inner_html()).lower()
            pinned = "pinned" in html
            results.append((href, pinned))
        except Exception:
            continue
    return results


async def open_profile_post_from_grid(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    *,
    target_post_url: str = "",
) -> Page:
    chosen_url = target_post_url
    shortcode = extract_shortcode(chosen_url) if chosen_url else ""
    max_scroll_attempts = 18 if target_post_url else 1
    target_anchor = None
    for _ in range(max_scroll_attempts):
        links = await visible_post_links(page)
        if not links:
            raise RuntimeError("No visible Instagram posts found on profile grid")

        if not chosen_url:
            pinned = [url for url, is_pinned in links if is_pinned]
            chosen_url = pinned[0] if pinned else links[0][0]
            shortcode = extract_shortcode(chosen_url)

        anchor = page.locator(f'a[href*="/{shortcode}/"]').first
        if await anchor.count():
            target_anchor = anchor
            break
        if not target_post_url:
            break
        await human.human_wheel_scroll(1400)
        await human.pause(650, 1100)

    if target_anchor is None:
        raise RuntimeError(f"Target post anchor not found for shortcode {shortcode}")

    logger.info("Opening post %s from profile grid", chosen_url)
    existing_pages = list(page.context.pages)
    await human.human_click(target_anchor)
    deadline = datetime.now(timezone.utc) + timedelta(seconds=15)
    while datetime.now(timezone.utc) < deadline:
        if extract_shortcode(page.url) == shortcode:
            await human.pause(650, 1200)
            return page
        for candidate in page.context.pages:
            if candidate in existing_pages:
                continue
            if extract_shortcode(candidate.url) == shortcode:
                await candidate.bring_to_front()
                await human.pause(650, 1200)
                return candidate
        await human.pause(300, 550)
    logger.info("Grid click did not produce an active post page; falling back to direct navigation %s", chosen_url)
    await page.goto(chosen_url, wait_until="domcontentloaded", timeout=60000)
    await human.pause(900, 1500)
    if extract_shortcode(page.url) != shortcode:
        raise RuntimeError(f"Profile post did not open for {chosen_url}")
    return page


async def has_next_post_navigation(page: Page) -> bool:
    buttons = page.locator("button")
    count = await buttons.count()
    viewport = page.viewport_size or {"width": 1920, "height": 1080}
    for index in range(count):
        button = buttons.nth(index)
        try:
            if not await button.is_visible():
                continue
            aria = (_compact(await button.get_attribute("aria-label") or "")).lower()
            text = _compact((await button.inner_text()) or "").lower()
            html = ((await button.inner_html()) or "").lower()
            if aria != "next" and "next" not in text and 'aria-label="next"' not in html and ">next<" not in html:
                continue
            box = await button.bounding_box()
            if not box:
                continue
            if box["x"] >= viewport["width"] * 0.55:
                return True
        except Exception:
            continue
    return False


async def reopen_post_modal(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    blogger_url: str,
    post_url: str,
) -> Page:
    target_shortcode = extract_shortcode(post_url)
    if target_shortcode and extract_shortcode(page.url) == target_shortcode:
        await human.pause(450, 800)
        return page

    logger.info("Reopening saved post from profile grid: %s", post_url)
    await ensure_profile_page(page, human, logger, blogger_url)
    for _ in range(3):
        try:
            return await open_profile_post_from_grid(
                page,
                human,
                logger,
                target_post_url=post_url,
            )
        except Exception:
            await human.human_wheel_scroll(900)
            await human.pause(500, 900)
    raise RuntimeError(f"Could not reopen saved post modal for {post_url}")


async def open_next_post_from_profile_grid(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    blogger_url: str,
    current_post_url: str,
) -> Page | None:
    current_shortcode = extract_shortcode(current_post_url)
    if not current_shortcode:
        return None
    await return_to_profile_from_post(page, human, logger, blogger_url)
    links = await visible_post_links(page)
    ordered = [url for url, _ in links]
    try:
        current_index = next(index for index, url in enumerate(ordered) if extract_shortcode(url) == current_shortcode)
    except StopIteration:
        return None
    if current_index + 1 >= len(ordered):
        return None
    next_url = ordered[current_index + 1]
    logger.info("Falling back to grid-based next post: %s", next_url)
    return await open_profile_post_from_grid(page, human, logger, target_post_url=next_url)


async def read_post_snapshot(page: Page, blogger_handle: str) -> PostSnapshot:
    data = await page.evaluate(
        """() => {
            const visible = (el) => !!(el && el.offsetParent);
            const toHandle = (href) => {
              if (!href) return "";
              const m = href.match(/^\\/([^/?#]+)\\/?$/);
              return m ? m[1] : "";
            };
            const authors = [];
            const headerLinks = Array.from(document.querySelectorAll('header a[href^="/"]'));
            for (const link of headerLinks) {
              if (!visible(link)) continue;
              const handle = toHandle(link.getAttribute('href'));
              if (!handle) continue;
              if (!authors.includes(handle)) authors.push(handle);
              if (authors.length >= 2) break;
            }

            let caption = "";
            const h1 = Array.from(document.querySelectorAll('h1[dir="auto"], h1')).find(visible);
            if (h1) {
              caption = (h1.innerText || "").trim();
            }
            if (!caption) {
              const article = document.querySelector('article') || document.querySelector('main');
              if (article) {
                const authorSet = new Set(authors.map(x => x.toLowerCase()));
                const items = Array.from(article.querySelectorAll('ul li')).filter(visible);
                let best = "";
                for (const item of items) {
                  const firstLink = Array.from(item.querySelectorAll('a[href^="/"]')).find(visible);
                  const handle = firstLink ? toHandle(firstLink.getAttribute('href')) : "";
                  if (!handle || !authorSet.has(handle.toLowerCase())) continue;
                  let text = (item.innerText || "").trim();
                  if (!text) continue;
                  text = text.replace(/\\bFollow\\b/gi, "").replace(/\\bReply\\b/gi, "").trim();
                  if (text.length > best.length) best = text;
                }
                caption = best;
              }
            }

            const timeEl = Array.from(document.querySelectorAll('time[datetime]')).find(visible);
            return {
              url: window.location.href,
              datetime: timeEl ? timeEl.getAttribute('datetime') || "" : "",
              caption,
              authors,
            };
        }"""
    )
    caption_text = _compact(str(data.get("caption") or ""))
    authors = [author for author in data.get("authors") or [] if author]
    current_path_handle = ""
    parsed = urlparse(page.url)
    path_parts = [part for part in parsed.path.split("/") if part]
    if len(path_parts) >= 3 and path_parts[1] in {"p", "reel"}:
        current_path_handle = path_parts[0]
    ignored_handles = {
        canonical_handle(blogger_handle),
        canonical_handle(current_path_handle),
        *(canonical_handle(author) for author in authors),
    }
    mentioned = []
    for handle in HANDLE_RE.findall(caption_text):
        normalized = canonical_handle(handle)
        if not normalized or normalized in ignored_handles:
            continue
        if normalized not in {canonical_handle(item) for item in mentioned}:
            mentioned.append(handle)

    ad_likelihood, ad_reasoning = classify_ad_likelihood(caption_text, mentioned)
    date_iso = str(data.get("datetime") or "")
    post_date = None
    if date_iso:
        try:
            post_date = datetime.fromisoformat(date_iso.replace("Z", "+00:00"))
        except ValueError:
            post_date = None

    return PostSnapshot(
        post_url=page.url,
        post_date=post_date,
        post_date_iso=post_date.isoformat() if post_date else date_iso,
        authors=authors,
        caption_text=caption_text,
        candidate_handles=mentioned,
        ad_likelihood=ad_likelihood,
        ad_reasoning=ad_reasoning,
    )


def classify_ad_likelihood(caption_text: str, candidate_handles: list[str]) -> tuple[str, str]:
    text = (caption_text or "").lower()
    score = 0
    reasons: list[str] = []
    if candidate_handles:
        score += 2
        reasons.append("есть явное @упоминание")
    for keyword in AD_KEYWORDS:
        if keyword in text:
            score += 1
            reasons.append(f"ключевое слово: {keyword}")
    if len(candidate_handles) >= 2:
        score += 1
        reasons.append("несколько @упоминаний")
    if score >= 4:
        return "high", "; ".join(reasons[:4]) or "сильные признаки рекламной интеграции"
    if score >= 2:
        return "medium", "; ".join(reasons[:4]) or "есть признаки интеграции"
    return "low", "; ".join(reasons[:4]) or "прямых признаков рекламы мало"


async def go_to_next_post(page: Page, human: Humanizer, logger: logging.Logger) -> bool:
    buttons = page.locator("button")
    count = await buttons.count()
    viewport = page.viewport_size or {"width": 1920, "height": 1080}
    preferred: tuple[Locator, float] | None = None
    best: tuple[Locator, float] | None = None
    for index in range(count):
        button = buttons.nth(index)
        try:
            if not await button.is_visible():
                continue
            aria = (_compact(await button.get_attribute("aria-label") or "")).lower()
            icon = button.locator('svg[aria-label="Next"], title')
            icon_text = _compact((await button.inner_text()) or "")
            has_next = False
            if await icon.count():
                icon_html = ((await button.inner_html()) or "").lower()
                has_next = 'aria-label="next"' in icon_html or ">next<" in icon_html
            if not has_next and "next" not in icon_text.lower() and aria != "next":
                continue
            box = await button.bounding_box()
            if not box:
                continue
            if box["x"] > viewport["width"] + 60:
                continue
            if aria == "next" and box["x"] >= viewport["width"] * 0.55:
                if preferred is None or box["x"] > preferred[1]:
                    preferred = (button, box["x"])
            if best is None or box["x"] > best[1]:
                best = (button, box["x"])
        except Exception:
            continue
    target = preferred or best
    if target is None:
        return False
    logger.info("Moving to next post in modal")
    current_shortcode = extract_shortcode(page.url)
    strategies = (
        ("human click", lambda: human.human_click(target[0])),
        ("force click", lambda: target[0].click(force=True)),
        ("dom click", lambda: target[0].evaluate("(el) => el.click()")),
    )
    for label, action in strategies:
        logger.info("Trying next-post navigation via %s", label)
        await action()
        deadline = datetime.now(timezone.utc) + timedelta(seconds=8)
        while datetime.now(timezone.utc) < deadline:
            next_shortcode = extract_shortcode(page.url)
            if next_shortcode and next_shortcode != current_shortcode:
                await human.pause(650, 1200)
                return True
            await human.pause(250, 500)
    return False


async def read_profile_overview(profile_page: Page) -> dict[str, str | int]:
    async def safe_meta_content(selector: str) -> str:
        locator = profile_page.locator(selector).first
        try:
            if await locator.count():
                return (await locator.get_attribute("content", timeout=5000)) or ""
        except Exception:
            return ""
        return ""

    meta_title = await safe_meta_content('meta[property="og:title"]')
    meta_description = await safe_meta_content('meta[name="description"]')

    header_text = ""
    header = profile_page.locator("header").first
    if await header.count():
        try:
            header_text = _compact(await header.inner_text(timeout=5000))
        except Exception:
            header_text = ""

    external_link = ""
    if await header.count():
        links = header.locator('a[href^="http"]')
        link_count = await links.count()
        for index in range(link_count):
            try:
                href = await links.nth(index).get_attribute("href", timeout=3000)
            except Exception:
                href = ""
            if href and "instagram.com" not in href and "threads.com" not in href:
                external_link = href
                break

    display_name = ""
    title_match = re.match(r"^(.*?)\s+\(@([A-Za-z0-9._]+)\)", meta_title or "")
    if title_match:
        display_name = _compact(title_match.group(1))

    followers_text = meta_description or ""
    bio = meta_description or header_text or _compact(await profile_page.title())
    category_label = ""
    header_lower = header_text.lower()
    for keyword in BRAND_KEYWORDS:
        if keyword in header_lower:
            category_label = keyword
            break

    followers_count = extract_followers_count(meta_description, header_text)
    return {
        "display_name": display_name,
        "bio": bio,
        "category_label": category_label,
        "followers_text": followers_text,
        "external_link": external_link,
        "header_text": header_text,
        "followers_count": followers_count,
    }


async def open_brand_profile_tab(
    context: BrowserContext,
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    *,
    handle: str,
    screenshots_dir: Path,
    source_blogger_handle: str,
    source_post: PostSnapshot,
) -> BrandAssessment:
    profile_url = f"https://www.instagram.com/{handle}/"
    brand_page = await context.new_page()
    screenshot_name = f"{source_blogger_handle}_{handle}_{extract_shortcode(source_post.post_url) or 'post'}.png"
    screenshot_path = screenshots_dir / screenshot_name
    try:
        await brand_page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
        await human.pause(1100, 1800)
        brand_human = Humanizer(brand_page, settings=human.settings)
        await dismiss_instagram_popups(brand_page, brand_human, logger)
        overview = await read_profile_overview(brand_page)
        display_name = str(overview["display_name"])
        bio = str(overview["bio"])
        category_label = str(overview["category_label"])
        followers_text = str(overview["followers_text"])
        external_link = str(overview["external_link"])

        brand_likelihood, confidence, reasoning, is_brand, niche, account_kind, outreach_fit = classify_brand_profile(
            handle=handle,
            display_name=display_name,
            bio=bio,
            category_label=category_label,
            external_link=external_link,
        )
        await capture_screenshot(brand_page, screenshot_path, logger)

        return BrandAssessment(
            handle=handle,
            profile_url=profile_url,
            is_brand=is_brand,
            account_kind=account_kind,
            outreach_fit=outreach_fit,
            brand_likelihood=brand_likelihood,
            ad_likelihood=source_post.ad_likelihood,
            niche=niche,
            confidence=confidence,
            reasoning=reasoning,
            display_name=display_name,
            bio=bio,
            category_label=category_label,
            followers_text=followers_text,
            external_link=external_link,
            screenshot_path=str(screenshot_path),
            source_posts=[source_post.post_url],
        )
    finally:
        try:
            await brand_page.close()
        except Exception:
            pass


def classify_brand_profile(
    *,
    handle: str,
    display_name: str,
    bio: str,
    category_label: str,
    external_link: str,
) -> tuple[str, str, str, bool, str, str, str]:
    haystack = _compact(" ".join([handle, display_name, bio, category_label])).lower()
    score = 0
    reasons: list[str] = []
    niche = detect_priority_niche(handle, display_name, bio, category_label)
    account_kind = "unclear"

    if external_link:
        score += 2
        reasons.append("???? ??????? ??????")
    for keyword in BRAND_KEYWORDS:
        if keyword in haystack:
            score += 1
            reasons.append(f"brand-keyword: {keyword}")
            if not niche:
                niche = keyword

    has_brand_keywords = any(keyword in haystack for keyword in BRAND_KEYWORDS)
    has_person_keywords = any(keyword in haystack for keyword in PERSON_KEYWORDS)
    has_service_keywords = any(keyword in haystack for keyword in SERVICE_KEYWORDS)
    has_project_keywords = any(keyword in haystack for keyword in PROJECT_KEYWORDS)

    if has_project_keywords:
        account_kind = "project_media"
    elif has_brand_keywords and not has_person_keywords:
        account_kind = "brand_store"
    elif has_service_keywords:
        account_kind = "service_provider"
    elif has_person_keywords:
        account_kind = "personal_creator"

    if has_person_keywords:
        score -= 2
        reasons.append("???????? ??? ?????? ??????? ??? ??????????")
    if re.fullmatch(r"[a-z]+[._]?[a-z]+[._]?[a-z0-9]*", handle.lower()) and score == 0:
        reasons.append("handle ??? ?? ???? ???????????")

    if account_kind == "brand_store":
        outreach_fit = "high" if score >= 2 else "medium"
    elif account_kind == "service_provider":
        outreach_fit = "medium"
    elif account_kind == "project_media":
        outreach_fit = "low"
    else:
        outreach_fit = "low"

    is_brand = account_kind in {"brand_store", "service_provider"}
    if not external_link and has_person_keywords and not has_brand_keywords and not has_service_keywords:
        return "low", "low", "; ".join(reasons[:4]) or "??????? ?????? ????? ?? ??????", False, niche, account_kind, outreach_fit

    if score >= 4:
        return "high", "high", "; ".join(reasons[:4]) or "??????? ????? ?? ?????", is_brand, niche, account_kind, outreach_fit
    if score >= 2:
        return "medium", "medium", "; ".join(reasons[:4]) or "???? ???????? ????????????? ????????", is_brand, niche, account_kind, outreach_fit
    return "low", "low", "; ".join(reasons[:4]) or "??????? ?????? ????? ?? ??????", is_brand and outreach_fit != "low", niche, account_kind, outreach_fit


def build_following_output_paths(job: dict, source_blogger_handle: str) -> tuple[Path, Path]:
    base_dir = Path(job["outputs"]["following_candidates_dir"])
    source_dir = base_dir / safe_handle_slug(source_blogger_handle)
    screenshots_dir = source_dir / "screenshots"
    return source_dir, screenshots_dir


def build_following_report_path(job: dict, source_blogger_handle: str) -> Path:
    source_dir, _ = build_following_output_paths(job, source_blogger_handle)
    return source_dir / "shortlist.md"


def build_following_brand_report_path(job: dict, source_blogger_handle: str) -> Path:
    source_dir, _ = build_following_output_paths(job, source_blogger_handle)
    return source_dir / "brands.md"


def build_following_global_report_path(job: dict) -> Path:
    base_dir = Path(job["outputs"]["following_candidates_dir"])
    return base_dir / "following_global.md"


def build_blogger_brand_report_path(job: dict, blogger_handle: str) -> Path:
    base_dir = Path(job["outputs"].get("brands_by_blogger_dir", "output/instagram_brand_search/brands/by_blogger"))
    return base_dir / safe_handle_slug(blogger_handle) / "collabs.md"


def build_brand_links_excel_path(job: dict) -> Path:
    brand_links_path = Path(job["outputs"]["discovered_brand_links_md"])
    return brand_links_path.with_suffix(".xlsx")


def build_run_exports_dir(job: dict) -> Path | None:
    run_dir = job.get("_run_dir")
    if not run_dir:
        return None
    return Path(str(run_dir)) / "exports"


def build_progress_output_paths(job: dict) -> tuple[Path, Path]:
    base_dir = Path(job["outputs"]["discovered_brand_links_md"]).parent.parent
    return base_dir / "run_status.md", base_dir / "run_status.json"


def _xlsx_column_name(index: int) -> str:
    result = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _xlsx_inline_cell(ref: str, value: object) -> str:
    if value is None:
        text = ""
    else:
        text = str(value)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{ref}"><v>{value}</v></c>'
    return f'<c r="{ref}" t="inlineStr"><is><t xml:space="preserve">{xml_escape(text)}</t></is></c>'


def _xlsx_clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)


def write_simple_xlsx(workbook_path: Path, sheets: list[tuple[str, list[list[object]]]]) -> None:
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        zebra_fill = PatternFill(fill_type="solid", fgColor="F7FBFF")
        thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))
        link_font = Font(color="0563C1", underline="single")

        for sheet_name, rows in sheets:
            worksheet = workbook.create_sheet(title=sheet_name[:31] or "Sheet1")
            for row in rows:
                worksheet.append(
                    [
                        value
                        if isinstance(value, (int, float, bool)) or value is None
                        else _xlsx_clean_text(value)
                        for value in row
                    ]
                )
            if not rows:
                continue

            max_cols = max(len(row) for row in rows)
            max_rows = len(rows)
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.zoomScale = 90
            worksheet.row_dimensions[1].height = 24

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            if max_rows >= 2 and max_cols >= 1:
                end_ref = f"{get_column_letter(max_cols)}{max_rows}"
                worksheet.auto_filter.ref = f"A1:{end_ref}"

            wrap_headers = {
                "Source Bloggers",
                "Run Labels",
                "Ad Reasoning",
                "Caption Excerpt",
                "Reasoning",
                "Bio",
                "External Link",
                "Latest Source Post",
                "Profile URL",
                "Post URL",
                "Source Workbook",
                "Structured Folder",
            }

            for col_idx in range(1, max_cols + 1):
                column_letter = get_column_letter(col_idx)
                header_value = str(worksheet.cell(row=1, column=col_idx).value or "")
                max_length = len(header_value)
                for row_idx in range(2, max_rows + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    text = "" if cell.value is None else str(cell.value)
                    if row_idx % 2 == 0:
                        cell.fill = zebra_fill
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=header_value in wrap_headers or len(text) > 60,
                    )
                    if text.startswith("http://") or text.startswith("https://"):
                        cell.hyperlink = text
                        cell.font = link_font
                    max_length = max(max_length, min(len(text), 80))

                if header_value in wrap_headers:
                    width = min(max(max_length, 18), 42)
                elif "count" in header_value.lower() or header_value.lower().endswith("run"):
                    width = min(max(max_length + 2, 10), 16)
                else:
                    width = min(max(max_length + 2, 12), 28)
                worksheet.column_dimensions[column_letter].width = width

        workbook.save(workbook_path)
        return
    except Exception:
        pass

    with zipfile.ZipFile(workbook_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
"""
            + "\n".join(
                f'  <Override PartName="/xl/worksheets/sheet{index}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
                for index in range(1, len(sheets) + 1)
            )
            + "\n</Types>\n",
        )
        archive.writestr(
            "_rels/.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>
""",
        )
        archive.writestr(
            "docProps/app.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Codex</Application>
</Properties>
""",
        )
        archive.writestr(
            "docProps/core.xml",
            f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" xmlns:dcmitype="http://purl.org/dc/dcmitype/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:creator>Codex</dc:creator>
  <cp:lastModifiedBy>Codex</cp:lastModifiedBy>
  <dcterms:created xsi:type="dcterms:W3CDTF">{datetime.now(timezone.utc).isoformat()}</dcterms:created>
</cp:coreProperties>
""",
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
"""
            + "\n".join(
                f'  <Relationship Id="rId{index}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{index}.xml"/>'
                for index in range(1, len(sheets) + 1)
            )
            + "\n</Relationships>\n",
        )
        archive.writestr(
            "xl/workbook.xml",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
"""
            + "\n".join(
                f'    <sheet name="{xml_escape(name[:31])}" sheetId="{index}" r:id="rId{index}"/>'
                for index, (name, _) in enumerate(sheets, start=1)
            )
            + "\n  </sheets>\n</workbook>\n",
        )
        for sheet_index, (_, rows) in enumerate(sheets, start=1):
            max_cols = max((len(row) for row in rows), default=1)
            sheet_rows: list[str] = []
            for row_index, row in enumerate(rows, start=1):
                cells = []
                for col_index in range(1, max_cols + 1):
                    value = row[col_index - 1] if col_index - 1 < len(row) else ""
                    cells.append(_xlsx_inline_cell(f"{_xlsx_column_name(col_index)}{row_index}", value))
                sheet_rows.append(f'<row r="{row_index}">{"".join(cells)}</row>')
            dimension_end = f"{_xlsx_column_name(max_cols)}{max(len(rows), 1)}"
            archive.writestr(
                f"xl/worksheets/sheet{sheet_index}.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
"""
                + f'  <dimension ref="A1:{dimension_end}"/>\n'
                + '  <sheetData>\n'
                + "\n".join(f"    {row}" for row in sheet_rows)
                + "\n  </sheetData>\n</worksheet>\n",
            )


def build_brand_links_excel_sheets(state: InstagramBrandSearchState) -> list[tuple[str, list[list[object]]]]:
    brand_rows: list[list[object]] = [[
        "Handle",
        "Profile URL",
        "Display Name",
        "Account Kind",
        "Outreach Fit",
        "Brand Likelihood",
        "Ad Likelihood",
        "Niche",
        "Category Label",
        "External Link",
        "Unique Bloggers Count",
        "Source Bloggers",
        "Source Posts Count",
        "Latest Source Post",
    ]]
    source_rows: list[list[object]] = [[
        "Brand Handle",
        "Blogger Handle",
        "Post Date",
        "Post URL",
        "Ad Likelihood",
        "Ad Reasoning",
        "Caption Excerpt",
    ]]
    blogger_rows: list[list[object]] = [[
        "Blogger Handle",
        "Profile URL",
        "Scanned Posts",
        "Candidate Mentions",
        "Accepted Brand Handles",
        "Stopped Due To Date",
        "Following Expansion Complete",
    ]]

    for handle, record in collect_exportable_brand_records(state):
        source_bloggers = sorted(
            {
                canonical_handle(str(source.get("blogger_handle", "")))
                for source in record.get("sources", [])
                if canonical_handle(str(source.get("blogger_handle", "")))
            }
        )
        latest_source = record.get("sources", [])[-1].get("post_url", "") if record.get("sources") else ""
        brand_rows.append([
            handle,
            record.get("profile_url", ""),
            normalize_text(record.get("display_name", "")),
            record.get("account_kind", ""),
            record.get("outreach_fit", ""),
            record.get("brand_likelihood", ""),
            record.get("ad_likelihood", ""),
            record.get("niche", ""),
            record.get("category_label", ""),
            record.get("external_link", ""),
            len(source_bloggers),
            ", ".join(source_bloggers),
            len(record.get("sources", [])),
            latest_source,
        ])
        for source in record.get("sources", []):
            source_rows.append([
                handle,
                source.get("blogger_handle", ""),
                source.get("post_date_iso", ""),
                source.get("post_url", ""),
                source.get("ad_likelihood", ""),
                normalize_text(source.get("ad_reasoning", "")),
                normalize_text(source.get("caption_excerpt", "")),
            ])

    for blogger_url, stats in sorted(state.blogger_stats.items()):
        blogger_handle = stats.get("handle") or extract_handle_from_url(blogger_url)
        blogger_rows.append([
            blogger_handle,
            blogger_url,
            int(stats.get("scanned_posts", 0) or 0),
            int(stats.get("candidate_mentions", 0) or 0),
            ", ".join(stats.get("accepted_brand_handles", []) or []),
            "yes" if stats.get("stopped_due_to_date") else "no",
            "yes" if blogger_url in state.completed_following_expansions else "no",
        ])

    return [
        ("Brands", brand_rows),
        ("Sources", source_rows),
        ("Blogger Summary", blogger_rows),
    ]


def write_brand_links_excel_outputs(job: dict, state: InstagramBrandSearchState) -> None:
    signature = (
        len(state.brand_records),
        sum(len(record.get("sources", [])) for record in state.brand_records.values()),
        len(state.blogger_stats),
    )
    if job.get("_brand_links_excel_signature") == signature:
        return

    workbook_path = build_brand_links_excel_path(job)
    sheets = build_brand_links_excel_sheets(state)
    write_simple_xlsx(workbook_path, sheets)

    run_exports_dir = build_run_exports_dir(job)
    if run_exports_dir is not None:
        run_exports_dir.mkdir(parents=True, exist_ok=True)
        write_simple_xlsx(run_exports_dir / "brand_links.xlsx", sheets)

    job["_brand_links_excel_signature"] = signature


def compute_run_progress(job: dict, state: InstagramBrandSearchState) -> dict:
    exportable_brand_records = collect_exportable_brand_records(state)
    seed_targets = load_blogger_targets(
        Path(job["inputs"]["blogger_list_file"]),
        limit=seed_target_limit(job),
    )
    seed_urls = [normalize_instagram_url(target.profile_url) for target in seed_targets]
    seed_set = set(seed_urls)
    completed_seed = [url for url in state.completed_bloggers if normalize_instagram_url(url) in seed_set]
    completed_following = [url for url in state.completed_following_expansions if normalize_instagram_url(url) in seed_set]
    following_enabled = bool(job.get("following_scan_policy", {}).get("enabled", True))
    selected_enabled = bool(job.get("following_scan_policy", {}).get("scan_selected_targets_after_discovery", True))
    following_targets = load_state_targets(state, job) if selected_enabled else []
    following_target_urls = [normalize_instagram_url(target.profile_url) for target in following_targets]
    following_target_completed = [url for url in following_target_urls if url in {normalize_instagram_url(item) for item in state.completed_bloggers}]

    phase = "completed"
    next_target_url = ""
    if len(completed_seed) < len(seed_urls):
        phase = "seed_scan"
        next_target_url = state.current_blogger_url or next(
            (url for url in seed_urls if url not in {normalize_instagram_url(item) for item in state.completed_bloggers}),
            "",
        )
    elif following_enabled and len(completed_following) < len(seed_urls):
        phase = "following_discovery"
        next_target_url = state.current_blogger_url or next(
            (url for url in seed_urls if url not in {normalize_instagram_url(item) for item in state.completed_following_expansions}),
            "",
        )
    elif selected_enabled and len(following_target_completed) < len(following_target_urls):
        phase = "following_targets_scan"
        next_target_url = state.current_blogger_url or next(
            (url for url in following_target_urls if url not in {normalize_instagram_url(item) for item in state.completed_bloggers}),
            "",
        )

    return {
        "run_label": job.get("_run_label", ""),
        "phase": phase,
        "current_blogger_url": state.current_blogger_url,
        "current_post_url": state.current_post_url,
        "current_post_date_iso": state.current_post_date_iso,
        "next_target_url": next_target_url,
        "seed_total": len(seed_urls),
        "seed_completed": len(completed_seed),
        "following_total": len(seed_urls) if following_enabled else 0,
        "following_completed": len(completed_following),
        "following_targets_total": len(following_target_urls),
        "following_targets_completed": len(following_target_completed),
        "brand_records": len(state.brand_records),
        "raw_brand_records": len(state.brand_records),
        "exportable_brand_records": len(exportable_brand_records),
        "blogger_stats": len(state.blogger_stats),
        "following_candidates": len(state.following_candidates),
    }


def write_run_progress_outputs(job: dict, state: InstagramBrandSearchState) -> None:
    progress = compute_run_progress(job, state)
    md_path, json_path = build_progress_output_paths(job)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")

    md_lines = [
        "# Instagram Brand Search Status",
        "",
        f"- Run: {progress.get('run_label') or 'current'}",
        f"- Phase: {progress.get('phase')}",
        f"- Current blogger: {progress.get('current_blogger_url') or 'none'}",
        f"- Current post: {progress.get('current_post_url') or 'none'}",
        f"- Next target: {progress.get('next_target_url') or 'none'}",
        f"- Seed progress: {progress.get('seed_completed', 0)} / {progress.get('seed_total', 0)}",
        f"- Following progress: {progress.get('following_completed', 0)} / {progress.get('following_total', 0)}",
        f"- Following-derived targets progress: {progress.get('following_targets_completed', 0)} / {progress.get('following_targets_total', 0)}",
        f"- Raw brand records: {progress.get('raw_brand_records', progress.get('brand_records', 0))}",
        f"- Exportable brand links: {progress.get('exportable_brand_records', 0)}",
        f"- Blogger stats: {progress.get('blogger_stats', 0)}",
        f"- Following candidates: {progress.get('following_candidates', 0)}",
        "",
    ]
    md_path.write_text("\n".join(md_lines), encoding="utf-8-sig")

    run_exports_dir = build_run_exports_dir(job)
    if run_exports_dir is not None:
        run_exports_dir.mkdir(parents=True, exist_ok=True)
        (run_exports_dir / "run_status.json").write_text(json.dumps(progress, ensure_ascii=False, indent=2), encoding="utf-8")
        (run_exports_dir / "run_status.md").write_text("\n".join(md_lines), encoding="utf-8-sig")


async def wait_for_following_dialog(page: Page, blogger_handle: str, timeout_ms: int = 12000) -> bool:
    try:
        await page.wait_for_function(
            """(handle) => {
                const dialogs = Array.from(document.querySelectorAll('div[role="dialog"]'));
                const dialog = dialogs[dialogs.length - 1];
                if (!dialog) return false;
                const text = (dialog.innerText || '').toLowerCase();
                const hasSearch = !!dialog.querySelector('input[placeholder*="Search"], input[aria-label*="Search"], input');
                const hasProfileLinks = Array.from(dialog.querySelectorAll('a[href^="/"]')).some((anchor) => {
                    const href = anchor.getAttribute('href') || '';
                    const match = href.match(/^\\/([A-Za-z0-9._]+)\\/?$/);
                    return !!match && match[1].toLowerCase() !== String(handle || '').toLowerCase();
                });
                return hasSearch || hasProfileLinks || text.includes('following') || text.includes('подпис');
            }""",
            blogger_handle,
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


async def open_following_dialog(page: Page, human: Humanizer, logger: logging.Logger, blogger_handle: str) -> None:
    patterns = (
        page.locator(f'a[href="/{blogger_handle}/following/"]').first,
        page.locator(f'a[href$="/{blogger_handle}/following/"]').first,
        page.locator('header a[href$="/following/"]').first,
        page.locator("a").filter(has_text=re.compile(r"following|подпис", re.IGNORECASE)).first,
        page.locator("button").filter(has_text=re.compile(r"following|подпис", re.IGNORECASE)).first,
    )
    for trigger in patterns:
        try:
            if await trigger.count() and await trigger.is_visible():
                logger.info("Opening following dialog for @%s", blogger_handle)
                await human.human_click(trigger)
                await page.locator('div[role="dialog"]').last.wait_for(state="visible", timeout=10000)
                await human.pause(700, 1200)
                return
        except Exception:
            continue

    try:
        direct_url = f"https://www.instagram.com/{blogger_handle}/following/"
        logger.info("Falling back to direct following URL for @%s -> %s", blogger_handle, direct_url)
        await page.goto(direct_url, wait_until="domcontentloaded", timeout=60000)
        await human.pause(1200, 1800)
        if "/following" in (page.url or ""):
            return
        dialog = page.locator('div[role="dialog"]').last
        await dialog.wait_for(state="visible", timeout=15000)
        return
    except Exception:
        pass
    raise RuntimeError(f"Could not open following dialog for @{blogger_handle}")


async def open_following_modal(page: Page, human: Humanizer, logger: logging.Logger, blogger_handle: str) -> None:
    if await wait_for_following_dialog(page, blogger_handle, timeout_ms=1200):
        return

    header = page.locator("header").first
    try:
        await header.wait_for(state="visible", timeout=10000)
        await page.locator(f'header a[href="/{blogger_handle}/following/"]').first.wait_for(state="visible", timeout=10000)
        await human.pause(900, 1500)
    except Exception:
        logger.info("Following trigger did not become visible in time for @%s", blogger_handle)

    patterns = (
        header.locator(f'a[href="/{blogger_handle}/following/"]').first,
        header.locator(f'a[href$="/{blogger_handle}/following/"]').first,
        header.locator('a[href$="/following/"]').filter(has_text=re.compile(r"(?:\\bfollowing\\b|РїРѕРґРїРёСЃ)", re.IGNORECASE)).first,
        header.locator("a").filter(has_text=re.compile(r"(?:\\bfollowing\\b|РїРѕРґРїРёСЃ)", re.IGNORECASE)).first,
        page.locator('header a[href$="/following/"]').first,
    )
    for trigger in patterns:
        try:
            if await trigger.count() and await trigger.is_visible():
                logger.info("Opening following modal for @%s", blogger_handle)
                await trigger.scroll_into_view_if_needed()
                for click_attempt in ("playwright_click", "dispatch_event", "human_click"):
                    try:
                        if click_attempt == "playwright_click":
                            await trigger.click(timeout=4000)
                        elif click_attempt == "dispatch_event":
                            await trigger.dispatch_event("click")
                        else:
                            await human.human_click(trigger)
                    except Exception:
                        continue
                    if await wait_for_following_dialog(page, blogger_handle, timeout_ms=5000):
                        await human.pause(1400, 2200)
                        return
        except Exception:
            continue

    try:
        clicked = await page.evaluate(
            """(handle) => {
                const header = document.querySelector('header');
                if (!header) return false;
                const normalizedHandle = String(handle || '').toLowerCase();
                const candidates = Array.from(header.querySelectorAll('a, button'));
                for (const node of candidates) {
                    const href = node.getAttribute('href') || '';
                    const text = (node.textContent || '').toLowerCase();
                    const looksLikeFollowingLink =
                        href.endsWith(`/${normalizedHandle}/following/`) ||
                        href.endsWith('/following/') ||
                        text.includes('following') ||
                        text.includes('РїРѕРґРїРёСЃ');
                    if (!looksLikeFollowingLink) continue;
                    node.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, composed: true }));
                    return true;
                }
                return false;
            }""",
            blogger_handle,
        )
        if clicked and await wait_for_following_dialog(page, blogger_handle):
            await human.pause(1400, 2200)
            return
    except Exception:
        pass
    raise RuntimeError(f"Could not open following modal for @{blogger_handle}")


async def read_following_dialog_handles(page: Page) -> tuple[list[str], bool, bool]:
    payload = await page.evaluate(
        """() => {
            const dialogs = Array.from(document.querySelectorAll('div[role="dialog"]'));
            const dialog = dialogs[dialogs.length - 1] || document.querySelector('main') || document.body;
            const handles = [];
            for (const anchor of dialog.querySelectorAll('a[href^="/"]')) {
              const href = anchor.getAttribute('href') || '';
              const match = href.match(/^\\/([A-Za-z0-9._]+)\\/?$/);
              if (!match) continue;
              const handle = match[1];
              if (!handles.includes(handle)) handles.push(handle);
            }
            const containers = [dialog, ...Array.from(dialog.querySelectorAll('div'))];
            let scrollBox = null;
            for (const el of containers) {
              if (el.scrollHeight > el.clientHeight + 120) {
                scrollBox = el;
                break;
              }
            }
            if (!scrollBox) {
              const doc = document.scrollingElement || document.documentElement || document.body;
              if (!doc) {
                return { handles, canScroll: false, atEnd: true };
              }
              const atEnd = doc.scrollTop + window.innerHeight >= doc.scrollHeight - 12;
              return { handles, canScroll: doc.scrollHeight > window.innerHeight + 120, atEnd };
            }
            const atEnd = scrollBox.scrollTop + scrollBox.clientHeight >= scrollBox.scrollHeight - 12;
            return { handles, canScroll: true, atEnd };
        }"""
    )
    return list(payload.get("handles", [])), bool(payload.get("canScroll")), bool(payload.get("atEnd"))


async def scroll_following_dialog(page: Page) -> None:
    await page.evaluate(
        """() => {
            const dialogs = Array.from(document.querySelectorAll('div[role="dialog"]'));
            const dialog = dialogs[dialogs.length - 1] || document.querySelector('main') || document.body;
            if (!dialog) return;
            const containers = [dialog, ...Array.from(dialog.querySelectorAll('div'))];
            let scrolled = false;
            for (const el of containers) {
              if (el.scrollHeight > el.clientHeight + 120) {
                el.scrollTop = Math.min(el.scrollHeight, el.scrollTop + Math.max(el.clientHeight * 0.85, 600));
                scrolled = true;
                break;
              }
            }
            if (!scrolled) {
              const doc = document.scrollingElement || document.documentElement || document.body;
              doc.scrollTop = Math.min(doc.scrollHeight, doc.scrollTop + Math.max(window.innerHeight * 0.85, 600));
            }
        }"""
    )


async def wait_for_following_handles(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    source_blogger_handle: str,
    *,
    max_attempts: int = 6,
) -> tuple[list[str], bool, bool]:
    latest: tuple[list[str], bool, bool] = ([], False, True)
    await human.pause(1300, 2200)
    for attempt in range(1, max_attempts + 1):
        handles, can_scroll, at_end = await read_following_dialog_handles(page)
        latest = (handles, can_scroll, at_end)
        logger.info(
            "Following dialog warmup for @%s: attempt=%s visible_handles=%s can_scroll=%s at_end=%s",
            source_blogger_handle,
            attempt,
            len(handles),
            can_scroll,
            at_end,
        )
        if handles:
            return latest
        if can_scroll and not at_end:
            await scroll_following_dialog(page)
        await human.pause(900, 1600)
    return latest


async def close_following_dialog(page: Page, human: Humanizer) -> None:
    try:
        await page.keyboard.press("Escape")
        await human.pause(350, 700)
    except Exception:
        pass


def upsert_following_candidate(state: InstagramBrandSearchState, candidate: FollowingCandidate) -> None:
    state.following_candidates[following_candidate_key(candidate.source_blogger_handle, candidate.handle)] = asdict(candidate)


def build_target_from_following_candidate(record: dict, job: dict | None = None) -> BloggerTarget | None:
    if job is not None:
        is_selected_target = is_selected_following_target_record(record, job)
    else:
        is_selected_target = bool(record.get("is_selected_target"))
    if not is_selected_target:
        return None
    profile_url = normalize_instagram_url(record.get("profile_url", ""))
    if not profile_url:
        return None
    handle = extract_handle_from_url(profile_url)
    return BloggerTarget(
        profile_url=profile_url,
        handle=handle,
        notes=f"following candidate of @{record.get('source_blogger_handle', '')}",
        source_kind="following_candidate",
        source_blogger_handle=str(record.get("source_blogger_handle", "")),
        source_blogger_url=str(record.get("source_blogger_url", "")),
    )


def is_following_record_qualified(record: dict, job: dict | None = None) -> bool:
    if job is not None:
        return bool(is_selected_following_target_record(record, job) or record.get("is_brand_like"))
    return bool(record.get("is_selected_target") or record.get("is_brand_like"))


def qualified_following_records_for_source(
    state: InstagramBrandSearchState,
    source_blogger_handle: str,
    job: dict | None = None,
) -> list[dict]:
    source_key = canonical_handle(source_blogger_handle)
    return [
        record
        for record in state.following_candidates.values()
        if canonical_handle(str(record.get("source_blogger_handle", ""))) == source_key and is_following_record_qualified(record, job)
    ]


async def inspect_following_candidate(
    context: BrowserContext,
    human: Humanizer,
    logger: logging.Logger,
    job: dict,
    *,
    source_blogger_handle: str,
    source_blogger_url: str,
    candidate_handle: str,
) -> FollowingCandidate:
    profile_url = f"https://www.instagram.com/{candidate_handle}/"
    candidate_page = await context.new_page()
    try:
        logger.info("Inspecting following candidate @%s from @%s", candidate_handle, source_blogger_handle)
        await human.pause(240, 520)
        await candidate_page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
        candidate_human = Humanizer(candidate_page, settings=human.settings)
        await candidate_human.pause(900, 1400)
        await dismiss_instagram_popups(candidate_page, candidate_human, logger)
        await candidate_human.pause(350, 700)
        try:
            await candidate_page.evaluate(
                """() => {
                    const amount = Math.min(Math.max(window.innerHeight * 0.18, 120), 220);
                    window.scrollBy(0, amount);
                }"""
            )
            await candidate_human.pause(220, 420)
            await candidate_page.evaluate("window.scrollTo(0, 0)")
            await candidate_human.pause(180, 360)
        except Exception:
            logger.info("Could not apply candidate review scroll for @%s", candidate_handle)
        overview = await read_profile_overview(candidate_page)
        await candidate_human.pause(180, 320)
        followers_count = int(overview["followers_count"])
        is_female_candidate, female_confidence, female_reasoning = classify_female_profile(
            handle=candidate_handle,
            display_name=str(overview["display_name"]),
            bio=str(overview["bio"]),
            category_label=str(overview["category_label"]),
        )
        is_brand_like, brand_confidence, brand_reasoning, matched_priority_niche = classify_following_brand_exclusion(
            handle=candidate_handle,
            display_name=str(overview["display_name"]),
            bio=str(overview["bio"]),
            category_label=str(overview["category_label"]),
            external_link=str(overview["external_link"]),
        )
        reject_brand_like = bool(job.get("following_scan_policy", {}).get("reject_brand_like_profiles", True))
        qualifies_followers_threshold = followers_within_selected_range(followers_count, job)
        is_selected_target = qualifies_followers_threshold and is_female_candidate and (not reject_brand_like or not is_brand_like)

        screenshot_path = ""
        if is_selected_target:
            _, screenshots_dir = build_following_output_paths(job, source_blogger_handle)
            screenshots_dir.mkdir(parents=True, exist_ok=True)
            target_path = screenshots_dir / f"{safe_handle_slug(candidate_handle)}.png"
            await capture_screenshot(candidate_page, target_path, logger)
            screenshot_path = str(target_path)

        return FollowingCandidate(
            source_blogger_handle=source_blogger_handle,
            source_blogger_url=source_blogger_url,
            handle=candidate_handle,
            profile_url=profile_url,
            display_name=str(overview["display_name"]),
            bio=str(overview["bio"]),
            category_label=str(overview["category_label"]),
            followers_text=str(overview["followers_text"]),
            followers_count=followers_count,
            external_link=str(overview["external_link"]),
            screenshot_path=screenshot_path,
            is_female_candidate=is_female_candidate,
            female_confidence=female_confidence,
            female_reasoning=female_reasoning,
            is_brand_like=is_brand_like,
            brand_confidence=brand_confidence,
            brand_reasoning=brand_reasoning,
            matched_priority_niche=matched_priority_niche,
            qualifies_followers_threshold=qualifies_followers_threshold,
            is_selected_target=is_selected_target,
        )
    finally:
        try:
            await candidate_page.close()
        except Exception:
            pass


async def fetch_instagram_profile_id(page: Page, handle: str) -> str:
    payload = await page.evaluate(
        """async ({ handle, appId }) => {
            const response = await fetch(`/api/v1/users/web_profile_info/?username=${encodeURIComponent(handle)}`, {
                method: 'GET',
                credentials: 'include',
                headers: { 'x-ig-app-id': appId },
            });
            const text = await response.text();
            let data = null;
            try {
                data = JSON.parse(text);
            } catch (error) {
                data = null;
            }
            return {
                ok: response.ok,
                status: response.status,
                userId: data?.data?.user?.id || '',
                preview: text.slice(0, 400),
            };
        }""",
        {"handle": handle, "appId": INSTAGRAM_WEB_APP_ID},
    )
    if payload.get("ok") and payload.get("userId"):
        return str(payload["userId"])

    profile_url = f"{INSTAGRAM_HOME_URL}{canonical_handle(handle)}/"
    await page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
    await page.wait_for_timeout(1500)
    html = await page.content()
    for pattern in (
        r'"profile_id":"(\d+)"',
        r'"target_id":"(\d+)"',
        r'"id":"(\d+)"',
    ):
        match = re.search(pattern, html)
        if match:
            return match.group(1)

    raise RuntimeError(f"Could not resolve Instagram profile id for @{handle}: status={payload.get('status')} body={payload.get('preview')}")


async def fetch_following_handles_via_api(
    page: Page,
    logger: logging.Logger,
    source_blogger_handle: str,
    *,
    max_profiles_to_review: int = 0,
) -> list[str]:
    user_id = await fetch_instagram_profile_id(page, source_blogger_handle)
    logger.info("Resolved Instagram user id for @%s -> %s", source_blogger_handle, user_id)

    discovered_handles: list[str] = []
    seen: set[str] = set()
    max_id = ""
    page_index = 0

    while True:
        page_index += 1
        payload = await page.evaluate(
            """async ({ userId, maxId, count, appId }) => {
                const url = new URL(`/api/v1/friendships/${userId}/following/`, window.location.origin);
                url.searchParams.set('count', String(count));
                if (maxId) url.searchParams.set('max_id', maxId);
                const response = await fetch(url.toString(), {
                    method: 'GET',
                    credentials: 'include',
                    headers: { 'x-ig-app-id': appId },
                });
                const text = await response.text();
                let data = null;
                try {
                    data = JSON.parse(text);
                } catch (error) {
                    data = null;
                }
                return {
                    ok: response.ok,
                    status: response.status,
                    preview: text.slice(0, 400),
                    nextMaxId: data?.next_max_id || '',
                    users: Array.isArray(data?.users)
                        ? data.users.map((user) => ({
                            username: user?.username || '',
                            fullName: user?.full_name || '',
                        }))
                        : [],
                };
            }""",
            {
                "userId": user_id,
                "maxId": max_id,
                "count": 50,
                "appId": INSTAGRAM_WEB_APP_ID,
            },
        )
        if not payload.get("ok"):
            raise RuntimeError(
                f"Could not fetch following for @{source_blogger_handle}: status={payload.get('status')} body={payload.get('preview')}"
            )

        page_users = payload.get("users", [])
        logger.info(
            "Following API page for @%s: page=%s users=%s next_max_id=%s",
            source_blogger_handle,
            page_index,
            len(page_users),
            "yes" if payload.get("nextMaxId") else "no",
        )
        for user in page_users:
            username = canonical_handle(str(user.get("username", "")))
            if not username or username == canonical_handle(source_blogger_handle) or username in seen:
                continue
            seen.add(username)
            discovered_handles.append(str(user.get("username", "")).strip())
            if max_profiles_to_review and len(discovered_handles) >= max_profiles_to_review:
                return discovered_handles

        max_id = str(payload.get("nextMaxId", "")).strip()
        if not max_id:
            break

    return discovered_handles


async def seed_allows_following_expansion(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    blogger: BloggerTarget,
) -> bool:
    if blogger.source_kind != "seed":
        return False

    await ensure_profile_loaded(page, human, logger, blogger.profile_url)
    overview = await read_profile_overview(page)
    is_brand_like, brand_confidence, brand_reasoning, matched_priority_niche = classify_following_brand_exclusion(
        handle=blogger.handle or extract_handle_from_url(blogger.profile_url),
        display_name=str(overview["display_name"]),
        bio=str(overview["bio"]),
        category_label=str(overview["category_label"]),
        external_link=str(overview["external_link"]),
    )
    if is_brand_like:
        logger.info(
            "Skipping following expansion for seed %s because it looks brand-like (%s, niche=%s): %s",
            blogger.profile_url,
            brand_confidence,
            matched_priority_niche or "unclear",
            brand_reasoning,
        )
        return False
    return True


async def discover_following_candidates(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    job: dict,
    *,
    state_path: Path,
    blogger: BloggerTarget,
) -> list[BloggerTarget]:
    source_blogger_handle = blogger.handle or extract_handle_from_url(blogger.profile_url)
    progress = state.following_progress_for(blogger.profile_url)
    await ensure_profile_loaded(page, human, logger, blogger.profile_url)

    policy = job.get("following_scan_policy", {})
    max_profiles_to_review = int(policy.get("max_profiles_to_review", 0) or 0)
    target_qualified_accounts = int(policy.get("target_qualified_accounts_per_seed", 20) or 0)
    discovered_handles = await fetch_following_handles_via_api(
        page,
        logger,
        source_blogger_handle,
        max_profiles_to_review=max_profiles_to_review,
    )
    progress["discovered_handles"] = discovered_handles
    progress["target_qualified_accounts"] = target_qualified_accounts
    if discovered_handles:
        progress["last_visible_handle"] = discovered_handles[-1]
    state.save(state_path)

    queued_targets: list[BloggerTarget] = []
    inspected_handles = set(progress.get("inspected_handles", []))
    qualified_handles = set(progress.get("qualified_handles", []))
    if not qualified_handles:
        qualified_handles = {
            str(record.get("handle", "")).strip()
            for record in qualified_following_records_for_source(state, source_blogger_handle, job)
            if str(record.get("handle", "")).strip()
        }
        progress["qualified_handles"] = list(qualified_handles)
    aborted_early = False
    for handle in discovered_handles:
        if target_qualified_accounts and len(qualified_handles) >= target_qualified_accounts:
            logger.info(
                "Following target reached for @%s: qualified=%s target=%s",
                source_blogger_handle,
                len(qualified_handles),
                target_qualified_accounts,
            )
            break
        if handle in inspected_handles:
            record = state.following_candidates.get(following_candidate_key(source_blogger_handle, handle))
            target = build_target_from_following_candidate(record, job) if record is not None else None
            if record is not None and is_following_record_qualified(record, job):
                qualified_handles.add(handle)
                progress["qualified_handles"] = list(qualified_handles)
            if target is not None:
                queued_targets.append(target)
            continue
        key = following_candidate_key(source_blogger_handle, handle)
        record = state.following_candidates.get(key)
        if record is None:
            try:
                candidate = await inspect_following_candidate(
                    page.context,
                    human,
                    logger,
                    job,
                    source_blogger_handle=source_blogger_handle,
                    source_blogger_url=blogger.profile_url,
                    candidate_handle=handle,
                )
            except Exception as exc:
                logger.info("Following candidate inspect failed for @%s from @%s: %s", handle, source_blogger_handle, exc)
                if "browser has been closed" in str(exc).lower() or "target page, context or browser has been closed" in str(exc).lower():
                    aborted_early = True
                    break
                continue
            upsert_following_candidate(state, candidate)
            record = asdict(candidate)
        inspected_handles.add(handle)
        progress["inspected_handles"] = list(inspected_handles)
        progress["last_processed_handle"] = handle
        if record is not None and is_following_record_qualified(record, job):
            qualified_handles.add(handle)
            progress["qualified_handles"] = list(qualified_handles)
        state.save(state_path)
        write_markdown_outputs(job, state)
        target = build_target_from_following_candidate(record, job)
        if target is not None:
            queued_targets.append(target)

    progress["list_exhausted"] = len(inspected_handles) >= len(discovered_handles)
    if not aborted_early and (
        len(inspected_handles) >= len(discovered_handles)
        or (target_qualified_accounts and len(qualified_handles) >= target_qualified_accounts)
    ):
        state.mark_following_expansion_completed(blogger.profile_url)
    state.save(state_path)
    return queued_targets


def upsert_brand_record(
    state: InstagramBrandSearchState,
    assessment: BrandAssessment,
    candidate: MentionCandidate,
) -> None:
    key = assessment.handle.lower()
    existing = state.brand_records.get(key)
    record = asdict(assessment)
    source_entry = {
        "blogger_handle": candidate.blogger_handle,
        "post_url": candidate.source_post_url,
        "post_date_iso": candidate.source_post_date_iso,
        "ad_likelihood": candidate.ad_likelihood,
        "ad_reasoning": candidate.ad_reasoning,
        "caption_excerpt": candidate.visible_context[:220],
    }
    if existing is None:
        record["sources"] = [source_entry]
        state.brand_records[key] = record
        return

    sources = existing.setdefault("sources", [])
    if source_entry not in sources:
        sources.append(source_entry)
    if assessment.screenshot_path:
        existing["screenshot_path"] = assessment.screenshot_path
    if assessment.external_link and not existing.get("external_link"):
        existing["external_link"] = assessment.external_link
    if assessment.account_kind and (existing.get("account_kind") in {"", "unclear", None}):
        existing["account_kind"] = assessment.account_kind
    if assessment.outreach_fit and assessment.outreach_fit != "low":
        existing["outreach_fit"] = assessment.outreach_fit
    if assessment.bio and len(assessment.bio) > len(existing.get("bio") or ""):
        existing["bio"] = assessment.bio
    if assessment.brand_likelihood == "high":
        existing["brand_likelihood"] = "high"
        existing["is_brand"] = existing.get("is_brand") or assessment.is_brand
    if assessment.ad_likelihood == "high":
        existing["ad_likelihood"] = "high"


def append_existing_brand_source(state: InstagramBrandSearchState, handle: str, candidate: MentionCandidate) -> None:
    existing = state.brand_records.get(handle.lower())
    if existing is None:
        return
    source_entry = {
        "blogger_handle": candidate.blogger_handle,
        "post_url": candidate.source_post_url,
        "post_date_iso": candidate.source_post_date_iso,
        "ad_likelihood": candidate.ad_likelihood,
        "ad_reasoning": candidate.ad_reasoning,
        "caption_excerpt": candidate.visible_context[:220],
    }
    sources = existing.setdefault("sources", [])
    if source_entry not in sources:
        sources.append(source_entry)
    source_posts = existing.setdefault("source_posts", [])
    if candidate.source_post_url and candidate.source_post_url not in source_posts:
        source_posts.append(candidate.source_post_url)


def update_blogger_stats(
    state: InstagramBrandSearchState,
    blogger_url: str,
    blogger_handle: str,
    *,
    scanned_increment: int = 0,
    candidate_increment: int = 0,
    accepted_handle: str = "",
    stopped_due_to_date: bool = False,
) -> None:
    stats = state.blogger_stats.setdefault(
        blogger_url,
        asdict(BloggerRunStats(profile_url=blogger_url, handle=blogger_handle)),
    )
    stats["handle"] = blogger_handle
    stats["scanned_posts"] = int(stats.get("scanned_posts", 0)) + scanned_increment
    stats["candidate_mentions"] = int(stats.get("candidate_mentions", 0)) + candidate_increment
    accepted = stats.setdefault("accepted_brand_handles", [])
    if accepted_handle and accepted_handle not in accepted:
        accepted.append(accepted_handle)
    if stopped_due_to_date:
        stats["stopped_due_to_date"] = True


def write_following_candidate_outputs(job: dict, state: InstagramBrandSearchState) -> None:
    min_followers, max_followers = get_following_selected_follower_bounds(job)
    follower_label = f"{min_followers // 1000}k-{max_followers // 1000}k" if max_followers else f"{min_followers // 1000}k+"
    grouped: dict[str, list[dict]] = {}
    for record in state.following_candidates.values():
        source_handle = str(record.get("source_blogger_handle", "")).strip()
        if not source_handle:
            continue
        grouped.setdefault(source_handle, []).append(record)

    for source_handle, records in grouped.items():
        source_dir, _ = build_following_output_paths(job, source_handle)
        source_dir.mkdir(parents=True, exist_ok=True)
        report_path = build_following_report_path(job, source_handle)
        brand_report_path = build_following_brand_report_path(job, source_handle)

        selected_records = dedupe_records_by_handle(
            [record for record in records if is_selected_following_target_record(record, job)]
        )
        selected_records.sort(key=lambda item: (-int(item.get("followers_count", 0) or 0), str(item.get("handle", ""))))
        brand_records = dedupe_records_by_handle(
            [record for record in records if record.get("is_brand_like") and is_exportable_brand_record(record)]
        )
        brand_records.sort(
            key=lambda item: (
                str(item.get("matched_priority_niche", "")),
                -int(item.get("followers_count", 0) or 0),
                str(item.get("handle", "")),
            )
        )
        rejected_brand_like = sum(1 for record in records if record.get("is_brand_like"))
        source_url = str(records[0].get("source_blogger_url", ""))
        progress = state.following_progress.get(source_url, {})
        qualified_total = len(progress.get("qualified_handles", []))
        target_qualified_total = int(progress.get("target_qualified_accounts", 0) or 0)

        lines = [f"# Shortlist @{source_handle}", ""]
        lines.append(f"- Исходный блогер: {md_link('@' + source_handle, source_url)}")
        lines.append(f"- Просмотрено профилей: {len(records)}")
        lines.append(f"- Подходящих профилей: {len(selected_records)}")
        lines.append(f"- Целевых аккаунтов всего: {qualified_total}" + (f" / {target_qualified_total}" if target_qualified_total else ""))
        lines.append(f"- Отсеяно как бренд/сервис: {rejected_brand_like}")
        lines.append(f"- Последний обработанный handle: @{progress.get('last_processed_handle', '')}" if progress.get("last_processed_handle") else "- Последний обработанный handle: none")
        lines.append("")

        if not selected_records:
            lines.append("Подходящих профилей пока нет.")
            lines.append("")
        else:
            for record in selected_records:
                screenshot_target = relative_markdown_target(report_path, str(record.get("screenshot_path", "")))
                lines.extend(
                    [
                        f"## @{record.get('handle', '')}",
                        f"- Метка: подписчица {source_handle}",
                        f"- Профиль: {md_link('@' + str(record.get('handle', '')), str(record.get('profile_url', '')))}",
                        f"- Скрин: {md_link('open screenshot', screenshot_target) if screenshot_target else ''}".rstrip(),
                        f"- Имя: {normalize_text(str(record.get('display_name', '')))}",
                        f"- Подписчики: {record.get('followers_count', 0)}",
                        f"- Followers raw: {normalize_text(str(record.get('followers_text', '')))}",
                        f"- Female match: {'yes' if record.get('is_female_candidate') else 'no'} ({record.get('female_confidence', '')})",
                        f"- Female reasoning: {normalize_text(str(record.get('female_reasoning', '')))}",
                        f"- Brand-like: {'yes' if record.get('is_brand_like') else 'no'} ({record.get('brand_confidence', '')})",
                        f"- Brand reasoning: {normalize_text(str(record.get('brand_reasoning', '')))}",
                        f"- Priority niche: {record.get('matched_priority_niche', '')}",
                        "",
                        "### Bio",
                        normalize_text(str(record.get("bio", ""))) or "none",
                        "",
                    ]
                )
        report_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8-sig")

        brand_lines = [f"# Brands In Following @{source_handle}", ""]
        brand_lines.append(f"- Исходный блогер: {md_link('@' + source_handle, source_url)}")
        brand_lines.append(f"- Найдено brand-like профилей: {len(brand_records)}")
        brand_lines.append(f"- Целевых аккаунтов всего: {qualified_total}" + (f" / {target_qualified_total}" if target_qualified_total else ""))
        brand_lines.append("")
        if not brand_records:
            brand_lines.append("Брендов в подписках пока не найдено.")
            brand_lines.append("")
        else:
            for record in brand_records:
                brand_lines.extend(
                    [
                        f"## @{record.get('handle', '')}",
                        f"- Профиль: {md_link('@' + str(record.get('handle', '')), str(record.get('profile_url', '')))}",
                        f"- Имя: {normalize_text(str(record.get('display_name', '')))}",
                        f"- Подписчики: {record.get('followers_count', 0)}",
                        f"- Followers raw: {normalize_text(str(record.get('followers_text', '')))}",
                        f"- Priority niche: {record.get('matched_priority_niche', '') or 'unclear'}",
                        f"- Brand-like: {'yes' if record.get('is_brand_like') else 'no'} ({record.get('brand_confidence', '')})",
                        f"- Brand reasoning: {normalize_text(str(record.get('brand_reasoning', '')))}",
                        "",
                        "### Bio",
                        normalize_text(str(record.get('bio', ''))) or "none",
                        "",
                    ]
                )
        brand_report_path.write_text("\n".join(brand_lines).strip() + "\n", encoding="utf-8-sig")

    global_report_path = build_following_global_report_path(job)
    global_report_path.parent.mkdir(parents=True, exist_ok=True)
    all_records = list(state.following_candidates.values())
    merged_selected = merge_following_records_by_handle(
        [record for record in all_records if is_selected_following_target_record(record, job)]
    )
    merged_selected.sort(
        key=lambda item: (
            -int(item.get("followers_count", 0) or 0),
            str(item.get("handle", "")),
        )
    )
    merged_brands = merge_following_records_by_handle(
        [record for record in all_records if record.get("is_brand_like") and is_exportable_brand_record(record)]
    )
    merged_brands.sort(
        key=lambda item: (
            str(item.get("matched_priority_niche", "")),
            -int(item.get("followers_count", 0) or 0),
            str(item.get("handle", "")),
        )
    )

    global_lines = ["# Following Global", ""]
    global_lines.append(f"- Уникальных shortlisted-профилей: {len(merged_selected)}")
    global_lines.append(f"- Уникальных brand-like профилей: {len(merged_brands)}")
    global_lines.append(f"- Всего записей второго этапа: {len(all_records)}")
    global_lines.append("")

    global_lines.append(f"## Shortlisted {follower_label}")
    global_lines.append("")
    if not merged_selected:
        global_lines.append("Подходящих shortlisted-профилей пока нет.")
        global_lines.append("")
    else:
        for record in merged_selected:
            source_handles = record.get("_source_handles", [])
            global_lines.extend(
                [
                    f"### @{record.get('handle', '')}",
                    f"- Профиль: {md_link('@' + str(record.get('handle', '')), str(record.get('profile_url', '')))}",
                    f"- Подписчики: {record.get('followers_count', 0)}",
                    f"- Источники: {', '.join('@' + handle for handle in source_handles) or 'none'}",
                    f"- Встречался у блогеров: {len(source_handles)}",
                    f"- Female match: {'yes' if record.get('is_female_candidate') else 'no'} ({record.get('female_confidence', '')})",
                    f"- Brand-like: {'yes' if record.get('is_brand_like') else 'no'} ({record.get('brand_confidence', '')})",
                    "",
                ]
            )

    global_lines.append("## Brands From Following")
    global_lines.append("")
    if not merged_brands:
        global_lines.append("Brand-like профилей в following пока нет.")
        global_lines.append("")
    else:
        for record in merged_brands:
            source_handles = record.get("_source_handles", [])
            global_lines.extend(
                [
                    f"### @{record.get('handle', '')}",
                    f"- Профиль: {md_link('@' + str(record.get('handle', '')), str(record.get('profile_url', '')))}",
                    f"- Имя: {normalize_text(str(record.get('display_name', '')))}",
                    f"- Подписчики: {record.get('followers_count', 0)}",
                    f"- Priority niche: {record.get('matched_priority_niche', '') or 'unclear'}",
                    f"- Источники: {', '.join('@' + handle for handle in source_handles) or 'none'}",
                    f"- Встречался у блогеров: {len(source_handles)}",
                    f"- Brand reasoning: {normalize_text(str(record.get('brand_reasoning', '')))}",
                    "",
                ]
            )

    global_report_path.write_text("\n".join(global_lines).strip() + "\n", encoding="utf-8-sig")

    phase1_md_path, phase1_txt_path = build_phase1_shortlist_paths(job)
    phase1_md_path.parent.mkdir(parents=True, exist_ok=True)
    previously_scanned = {extract_handle_from_url(url) for url in state.completed_bloggers}

    phase1_md_lines = ["# Shortlisted Bloggers For Phase 1", ""]
    phase1_md_lines.append(f"- Total shortlisted bloggers in follower range: {len(merged_selected)}")
    phase1_md_lines.append(f"- Follower range: {min_followers} to {max_followers if max_followers else 'unbounded'}")
    phase1_md_lines.append("- Source: following discovery shortlist, deduplicated across all seed bloggers")
    phase1_md_lines.append("")

    phase1_txt_lines: list[str] = []
    for record in merged_selected:
        handle = str(record.get("handle", "")).strip()
        profile_url = normalize_instagram_url(str(record.get("profile_url", "")))
        if not handle or not profile_url:
            continue
        source_handles = record.get("_source_handles", [])
        phase1_md_lines.extend(
            [
                f"## @{handle}",
                f"- Profile: {md_link('@' + handle, profile_url)}",
                f"- Followers: {int(record.get('followers_count', 0) or 0)}",
                f"- Female match: {'yes' if record.get('is_female_candidate') else 'no'} ({record.get('female_confidence', '')})",
                f"- Found via seed bloggers: {', '.join('@' + source_handle for source_handle in source_handles) or 'none'}",
                f"- Seen in sources: {len(source_handles)}",
                f"- Already scanned in phase 1 before: {'yes' if handle in previously_scanned else 'no'}",
                f"- Raw followers text: {normalize_text(str(record.get('followers_text', '')))}",
                f"- Bio: {normalize_text(str(record.get('bio', ''))) or 'none'}",
                "",
            ]
        )
        phase1_txt_lines.append(profile_url)

    if not phase1_txt_lines:
        phase1_md_lines.append("No shortlisted bloggers in the configured follower range yet.")
        phase1_md_lines.append("")

    phase1_md_path.write_text("\n".join(phase1_md_lines).strip() + "\n", encoding="utf-8-sig")
    phase1_txt_path.write_text("\n".join(phase1_txt_lines).strip() + ("\n" if phase1_txt_lines else ""), encoding="utf-8-sig")


def write_markdown_outputs(job: dict, state: InstagramBrandSearchState) -> None:
    outputs = job["outputs"]
    blogger_summary_path = Path(outputs["blogger_summary_md"])
    brand_links_path = Path(outputs["discovered_brand_links_md"])
    brand_dossiers_path = Path(outputs["extracted_candidates_md"])
    exportable_brand_records = collect_exportable_brand_records(state)
    for path in (blogger_summary_path, brand_links_path, brand_dossiers_path):
        path.parent.mkdir(parents=True, exist_ok=True)

    blogger_lines = ["# Blogger Summary", ""]
    for blogger_url, stats in sorted(state.blogger_stats.items()):
        blogger_handle = stats.get("handle") or extract_handle_from_url(blogger_url)
        blogger_report_path = build_blogger_brand_report_path(job, blogger_handle)
        blogger_report_path.parent.mkdir(parents=True, exist_ok=True)
        blogger_brand_records = [
            (handle, record)
            for handle, record in exportable_brand_records
            if any(source.get("blogger_handle") == blogger_handle for source in record.get("sources", []))
        ]
        accepted_handles = sorted(
            handle for handle, _ in blogger_brand_records
        )
        blogger_lines.extend(
            [
                f"## {blogger_handle}",
                f"- Profile: {md_link('@' + blogger_handle, blogger_url)}",
                f"- Scanned posts: {stats.get('scanned_posts', 0)}",
                f"- Candidate mentions: {stats.get('candidate_mentions', 0)}",
                f"- Accepted brand-like handles: {', '.join(accepted_handles) or 'none'}",
                f"- Stopped due to 1y cutoff: {'yes' if stats.get('stopped_due_to_date') else 'no'}",
                f"- Following expansion complete: {'yes' if blogger_url in state.completed_following_expansions else 'no'}",
                "",
            ]
        )
        blogger_lines_local = [f"# Collabs @{blogger_handle}", ""]
        blogger_lines_local.extend(
            [
                f"- Profile: {md_link('@' + blogger_handle, blogger_url)}",
                f"- Scanned posts: {stats.get('scanned_posts', 0)}",
                f"- Candidate mentions: {stats.get('candidate_mentions', 0)}",
                f"- Accepted brand-like handles: {', '.join(accepted_handles) or 'none'}",
                f"- Following expansion complete: {'yes' if blogger_url in state.completed_following_expansions else 'no'}",
                "",
            ]
        )
        if not blogger_brand_records:
            blogger_lines_local.append("Подходящих brand-like аккаунтов пока нет.")
            blogger_lines_local.append("")
        else:
            for handle, record in blogger_brand_records:
                screenshot_target = relative_markdown_target(blogger_report_path, record.get("screenshot_path", ""))
                latest_source = ""
                if record.get("sources"):
                    source_posts = [source.get("post_url", "") for source in record.get("sources", []) if source.get("blogger_handle") == blogger_handle]
                    latest_source = source_posts[-1] if source_posts else ""
                blogger_lines_local.extend(
                    [
                        f"## @{handle}",
                        f"- Profile: {md_link('@' + handle, record.get('profile_url', ''))}",
                        f"- Screenshot: {md_link('open screenshot', screenshot_target) if screenshot_target else ''}".rstrip(),
                        f"- Account kind: {record.get('account_kind', '')}",
                        f"- Outreach fit: {record.get('outreach_fit', '')}",
                        f"- Brand likelihood: {record.get('brand_likelihood', '')}",
                        f"- Ad likelihood: {record.get('ad_likelihood', '')}",
                        f"- Latest source post: {md_link('open post', latest_source) if latest_source else ''}".rstrip(),
                        "",
                    ]
                )
        blogger_report_path.write_text("\n".join(blogger_lines_local).strip() + "\n", encoding="utf-8-sig")
    blogger_summary_path.write_text("\n".join(blogger_lines).strip() + "\n", encoding="utf-8-sig")

    link_lines = ["# Brand Links", ""]
    for handle, record in exportable_brand_records:
        screenshot_target = relative_markdown_target(brand_links_path, record.get("screenshot_path", ""))
        latest_source = ""
        if record.get("sources"):
            latest_source = record["sources"][-1].get("post_url", "")
        source_bloggers = sorted(
            {
                canonical_handle(str(source.get("blogger_handle", "")))
                for source in record.get("sources", [])
                if canonical_handle(str(source.get("blogger_handle", "")))
            }
        )
        link_lines.extend(
            [
                f"## @{handle}",
                f"- Profile: {md_link('@' + handle, record.get('profile_url', ''))}",
                f"- Screenshot: {md_link('open screenshot', screenshot_target) if screenshot_target else ''}".rstrip(),
                f"- Account kind: {record.get('account_kind', '')}",
                f"- Outreach fit: {record.get('outreach_fit', '')}",
                f"- Brand likelihood: {record.get('brand_likelihood')}",
                f"- Ad likelihood: {record.get('ad_likelihood')}",
                f"- Source bloggers: {', '.join(source_bloggers) or 'none'}",
                f"- Sources: {len(record.get('sources', []))}",
                f"- Latest source post: {md_link('open post', latest_source) if latest_source else ''}".rstrip(),
                "",
            ]
        )
    brand_links_path.write_text("\n".join(link_lines).strip() + "\n", encoding="utf-8-sig")

    dossier_lines = ["# Brand Dossiers", ""]
    for handle, record in exportable_brand_records:
        screenshot_target = relative_markdown_target(brand_dossiers_path, record.get("screenshot_path", ""))
        bio = normalize_text(record.get("bio", ""))
        reasoning = normalize_text(record.get("reasoning", ""))
        display_name = normalize_text(record.get("display_name", ""))
        dossier_lines.extend(
            [
                f"## @{handle}",
                f"- Profile: {md_link('@' + handle, record.get('profile_url', ''))}",
                f"- Screenshot: {md_link('open screenshot', screenshot_target) if screenshot_target else ''}".rstrip(),
                f"- Brand likelihood: {record.get('brand_likelihood', '')}",
                f"- Ad likelihood: {record.get('ad_likelihood', '')}",
                f"- Is brand: {'yes' if record.get('is_brand') else 'no'}",
                f"- Account kind: {record.get('account_kind', '')}",
                f"- Outreach fit: {record.get('outreach_fit', '')}",
                f"- Display name: {display_name}",
                f"- Niche: {record.get('niche', '')}",
                f"- Category label: {record.get('category_label', '')}",
                f"- External link: {md_link('open external link', record.get('external_link', '')) if record.get('external_link') else ''}".rstrip(),
                f"- Reasoning: {reasoning}",
                "",
                "### Bio",
                bio or "none",
                "",
                "### Source Posts",
            ]
        )
        sources = record.get("sources", [])
        if not sources:
            dossier_lines.append("- none")
        else:
            for source in sources:
                post_url = source.get("post_url", "")
                excerpt = normalize_text(source.get("caption_excerpt", ""))
                dossier_lines.append(
                    f"- {source.get('blogger_handle', '')} | {source.get('post_date_iso', '')} | {md_link('post', post_url) if post_url else ''} | ad={source.get('ad_likelihood', '')} | {normalize_text(source.get('ad_reasoning', ''))}"
                )
                if excerpt:
                    dossier_lines.append(f"  excerpt: {excerpt}")
        dossier_lines.append("")
    brand_dossiers_path.write_text("\n".join(dossier_lines).strip() + "\n", encoding="utf-8-sig")
    write_following_candidate_outputs(job, state)
    write_brand_links_excel_outputs(job, state)
    write_run_progress_outputs(job, state)


def is_post_older_than_window(post_date: datetime | None, target_days: int) -> bool:
    if post_date is None:
        return False
    if target_days <= 0:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(days=target_days)
    current = post_date.astimezone(timezone.utc) if post_date.tzinfo else post_date.replace(tzinfo=timezone.utc)
    return current < cutoff


def update_current_post_pointer(state: InstagramBrandSearchState, blogger_url: str, post_url: str) -> None:
    checkpoint = state.checkpoint_for(blogger_url)
    checkpoint.current_post_url = post_url
    checkpoint.last_processed_shortcode = extract_shortcode(post_url)
    checkpoint.current_post_date_iso = ""
    state.current_post_url = post_url
    state.current_post_date_iso = ""


async def process_current_post(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    job: dict,
    screenshots_dir: Path,
    blogger_url: str,
    blogger_handle: str,
) -> tuple[str, list[BrandAssessment]]:
    checkpoint = state.checkpoint_for(blogger_url)
    snapshot = await read_post_snapshot(page, blogger_handle)
    state.current_post_url = snapshot.post_url
    state.current_post_date_iso = snapshot.post_date_iso
    checkpoint.current_post_url = snapshot.post_url
    checkpoint.current_post_date_iso = snapshot.post_date_iso
    checkpoint.last_processed_shortcode = extract_shortcode(snapshot.post_url)

    if is_post_older_than_window(snapshot.post_date, job["scan_policy"]["target_period_days"]):
        existing_scanned = int(state.blogger_stats.get(blogger_url, {}).get("scanned_posts", 0))
        if existing_scanned > 0:
            update_blogger_stats(state, blogger_url, blogger_handle, stopped_due_to_date=True)
            return "stop", []
        return "skip_old", []

    if snapshot.post_url not in checkpoint.processed_post_urls:
        checkpoint.processed_post_urls.append(snapshot.post_url)
    checkpoint.processed_posts_count += 1
    update_blogger_stats(
        state,
        blogger_url,
        blogger_handle,
        scanned_increment=1,
        candidate_increment=len(snapshot.candidate_handles),
    )

    assessments: list[BrandAssessment] = []
    for handle in snapshot.candidate_handles[: job["scan_policy"].get("max_candidates_per_post", 10)]:
        key = candidate_key(snapshot.post_url, handle)
        if key in checkpoint.processed_candidate_keys:
            continue
        checkpoint.processed_candidate_keys.append(key)
        candidate = MentionCandidate(
            blogger_handle=blogger_handle,
            source_post_url=snapshot.post_url,
            source_post_date_iso=snapshot.post_date_iso,
            source_authors=snapshot.authors,
            candidate_handle=handle,
            visible_context=snapshot.caption_text,
            caption_text=snapshot.caption_text,
            source_type="caption",
            ad_likelihood=snapshot.ad_likelihood,
            ad_reasoning=snapshot.ad_reasoning,
        )
        existing = state.brand_records.get(handle.lower())
        if existing is not None:
            logger.info("Skipping duplicate known handle @%s on %s", handle, snapshot.post_url)
            append_existing_brand_source(state, handle, candidate)
            continue
        try:
            assessment = await open_brand_profile_tab(
                page.context,
                page,
                human,
                logger,
                handle=handle,
                screenshots_dir=screenshots_dir,
                source_blogger_handle=blogger_handle,
                source_post=snapshot,
            )
        except Exception as exc:
            logger.info("Brand tab failed for @%s on %s: %s", handle, snapshot.post_url, exc)
            await page.bring_to_front()
            await human.pause(450, 850)
            continue
        upsert_brand_record(state, assessment, candidate)
        if assessment.is_brand:
            update_blogger_stats(state, blogger_url, blogger_handle, accepted_handle=handle)
        assessments.append(assessment)
        await page.bring_to_front()
        await human.pause(450, 850)

    return "continue", assessments


async def recover_modal_if_needed(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    blogger_url: str,
) -> Page:
    checkpoint = state.checkpoint_for(blogger_url)
    current_shortcode = extract_shortcode(page.url)
    target_shortcode = extract_shortcode(checkpoint.current_post_url or state.current_post_url)
    if target_shortcode and current_shortcode != target_shortcode:
        logger.info("Recovering modal state for %s", checkpoint.current_post_url)
        return await reopen_post_modal(page, human, logger, blogger_url, checkpoint.current_post_url)
    return page


async def run_blogger_scan(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    job: dict,
    state_path: Path,
    screenshots_dir: Path,
    blogger: BloggerTarget,
) -> Page:
    checkpoint = state.checkpoint_for(blogger.profile_url)
    state.current_blogger_url = blogger.profile_url
    await ensure_profile_page(page, human, logger, blogger.profile_url)
    blogger_handle = blogger.handle or extract_handle_from_url(blogger.profile_url)

    try:
        if checkpoint.current_post_url:
            page = await reopen_post_modal(page, human, logger, blogger.profile_url, checkpoint.current_post_url)
        else:
            page = await open_profile_post_from_grid(page, human, logger)
    except Exception as exc:
        logger.info("Skipping blogger %s because no accessible post grid was found: %s", blogger.profile_url, exc)
        state.mark_blogger_completed(blogger.profile_url)
        state.save(state_path)
        write_markdown_outputs(job, state)
        return page

    update_current_post_pointer(state, blogger.profile_url, page.url)

    fallback_limit = int(job["scan_policy"]["fallback_max_posts_per_blogger"])
    while checkpoint.processed_posts_count < fallback_limit:
        page = await recover_modal_if_needed(page, human, logger, state, blogger.profile_url)
        update_current_post_pointer(state, blogger.profile_url, page.url)
        action, _ = await process_current_post(
            page,
            human,
            logger,
            state,
            job,
            screenshots_dir,
            blogger.profile_url,
            blogger_handle,
        )
        state.save(state_path)
        write_markdown_outputs(job, state)
        if checkpoint.processed_posts_count >= fallback_limit:
            logger.info(
                "Reached scan limit (%s posts) for blogger %s",
                fallback_limit,
                blogger_handle,
            )
            break
        if action == "stop":
            break
        if action == "skip_old":
            try:
                moved = await go_to_next_post(page, human, logger)
            except Exception as exc:
                logger.info("Stopping blogger %s after next-post navigation failed during old-post skip: %s", blogger_handle, exc)
                break
            if moved:
                update_current_post_pointer(state, blogger.profile_url, page.url)
                state.save(state_path)
                write_markdown_outputs(job, state)
            else:
                logger.info("No further Next button found for blogger %s while skipping old pinned content", blogger_handle)
                break
            continue
        try:
            moved = await go_to_next_post(page, human, logger)
        except Exception as exc:
            logger.info("Stopping blogger %s after next-post navigation failed: %s", blogger_handle, exc)
            break
        if moved:
            update_current_post_pointer(state, blogger.profile_url, page.url)
            state.save(state_path)
            write_markdown_outputs(job, state)
        else:
            logger.info("No further Next button found for blogger %s", blogger_handle)
            break

    state.mark_blogger_completed(blogger.profile_url)
    state.save(state_path)
    write_markdown_outputs(job, state)
    return page


def load_state_targets(state: InstagramBrandSearchState, job: dict | None = None) -> list[BloggerTarget]:
    targets: list[BloggerTarget] = []
    for record in state.following_candidates.values():
        target = build_target_from_following_candidate(record, job)
        if target is not None:
            targets.append(target)
    return targets


def rotate_targets_for_resume(targets: list[BloggerTarget], current_blogger_url: str) -> list[BloggerTarget]:
    if not current_blogger_url:
        return targets
    start_index = next(
        (index for index, target in enumerate(targets) if normalize_instagram_url(target.profile_url) == normalize_instagram_url(current_blogger_url)),
        0,
    )
    return targets[start_index:] + targets[:start_index]


async def run_instagram_brand_search(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    job: dict,
    *,
    state_path: Path,
) -> None:
    screenshots_dir = Path(job["outputs"]["candidate_screenshots_dir"])
    screenshots_dir.mkdir(parents=True, exist_ok=True)

    seed_targets = load_blogger_targets(
        Path(job["inputs"]["blogger_list_file"]),
        limit=seed_target_limit(job),
    )
    seed_targets = rotate_targets_for_resume(seed_targets, state.current_blogger_url)

    for blogger in seed_targets:
        if blogger.profile_url not in state.completed_bloggers:
            logger.info("Scanning seed account %s", blogger.profile_url)
            page = await run_blogger_scan(page, human, logger, state, job, state_path, screenshots_dir, blogger)
            state.save(state_path)
            write_markdown_outputs(job, state)

    for blogger in seed_targets:
        should_expand_following = job.get("following_scan_policy", {}).get("enabled", True)
        if should_expand_following and blogger.profile_url not in state.completed_following_expansions:
            allows_following = await seed_allows_following_expansion(page, human, logger, blogger)
            if allows_following:
                logger.info("Expanding following list for %s after completing all seed scans", blogger.profile_url)
                await discover_following_candidates(
                    page,
                    human,
                    logger,
                    state,
                    job,
                    state_path=state_path,
                    blogger=blogger,
                )
                state.save(state_path)
                write_markdown_outputs(job, state)
            else:
                logger.info("Marking following expansion as completed/skipped for %s", blogger.profile_url)
                state.mark_following_expansion_completed(blogger.profile_url)
                state.save(state_path)
                write_markdown_outputs(job, state)

    if job.get("following_scan_policy", {}).get("scan_selected_targets_after_discovery", True):
        following_targets = load_state_targets(state, job)
        following_targets = rotate_targets_for_resume(following_targets, state.current_blogger_url)
        for blogger in following_targets:
            if blogger.profile_url in state.completed_bloggers:
                continue
            logger.info("Scanning following-derived account %s", blogger.profile_url)
            page = await run_blogger_scan(page, human, logger, state, job, state_path, screenshots_dir, blogger)
            state.save(state_path)
            write_markdown_outputs(job, state)

    write_markdown_outputs(job, state)


async def run_instagram_following_discovery(
    page: Page,
    human: Humanizer,
    logger: logging.Logger,
    state: InstagramBrandSearchState,
    job: dict,
    *,
    state_path: Path,
    force_rescan: bool = False,
) -> None:
    targets = load_blogger_targets(
        Path(job["inputs"]["blogger_list_file"]),
        limit=seed_target_limit(job),
    )
    for blogger in targets:
        if blogger.source_kind != "seed":
            logger.info("Skipping following-only seed %s because it is pre-classified as brand-like", blogger.profile_url)
            continue
        already_done = blogger.profile_url in state.completed_following_expansions
        if already_done and not force_rescan:
            logger.info("Skipping following expansion for %s because it is already completed", blogger.profile_url)
            continue
        if already_done and force_rescan:
            source_handle = blogger.handle or extract_handle_from_url(blogger.profile_url)
            state.completed_following_expansions = [
                url for url in state.completed_following_expansions if normalize_instagram_url(url) != normalize_instagram_url(blogger.profile_url)
            ]
            state.following_progress.pop(blogger.profile_url, None)
            state.following_candidates = {
                key: value
                for key, value in state.following_candidates.items()
                if canonical_handle(str(value.get("source_blogger_handle", ""))) != canonical_handle(source_handle)
            }
            source_dir, _ = build_following_output_paths(job, source_handle)
            if source_dir.exists():
                shutil.rmtree(source_dir, ignore_errors=True)
        allows_following = await seed_allows_following_expansion(page, human, logger, blogger)
        if not allows_following:
            state.mark_following_expansion_completed(blogger.profile_url)
            state.save(state_path)
            write_markdown_outputs(job, state)
            continue
        logger.info("Following-only scan for %s", blogger.profile_url)
        await discover_following_candidates(
            page,
            human,
            logger,
            state,
            job,
            state_path=state_path,
            blogger=blogger,
        )
        state.save(state_path)
        write_markdown_outputs(job, state)

    write_markdown_outputs(job, state)
