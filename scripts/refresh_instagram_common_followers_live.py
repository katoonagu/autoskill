from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeoutError, async_playwright

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from automation.adspower import AdsPowerClient
from automation.artifacts import setup_run_artifacts
from automation.browser import capture_screenshot, connect_profile
from automation.config import AdsPowerSettings
from automation.modules.instagram_brand_search.recipe import (
    Humanizer,
    build_humanizer,
    canonical_handle,
    dismiss_instagram_popups,
    parse_compact_number,
    read_profile_overview,
)


WORKBOOK_PATH = PROJECT_ROOT / "output" / "instagram_brand_search" / "brands" / "tables" / "brand_links_common.xlsx"
LIVE_CACHE_PATH = PROJECT_ROOT / "output" / "instagram_brand_search" / "brands" / "tables" / "live_profile_counts.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh Instagram follower counts in brand_links_common.xlsx")
    parser.add_argument("--limit", type=int, default=0, help="Maximum number of profiles to refresh")
    parser.add_argument("--profile-no", default="", help="Override AdsPower profile number")
    parser.add_argument("--handles", nargs="*", default=[], help="Optional specific brand handles to refresh")
    parser.add_argument("--all", action="store_true", help="Refresh all rows, not only rows with zero followers")
    return parser.parse_args()


def _extract_metric(raw_text: str, keywords: tuple[str, ...]) -> int:
    patterns = [
        re.compile(rf"([\d][\d\s.,]*\s*(?:k|m|b|тыс\.?|млн|млрд)?)\s*(?:{'|'.join(keywords)})", re.IGNORECASE),
        re.compile(rf"(?:{'|'.join(keywords)})\s*[:\-]?\s*([\d][\d\s.,]*\s*(?:k|m|b|тыс\.?|млн|млрд)?)", re.IGNORECASE),
    ]
    for pattern in patterns:
        match = pattern.search(raw_text)
        if match:
            return parse_compact_number(match.group(1))
    return 0


def _extract_profile_counts(*texts: str) -> dict[str, int]:
    raw_text = " ".join(str(text or "") for text in texts)
    return {
        "followers": _extract_metric(raw_text, ("followers?", "подписчик[а-я]*")),
        "following": _extract_metric(raw_text, ("following", "подписок", "подписки", "подписан")),
        "posts": _extract_metric(raw_text, ("posts?", "публикац(?:ий|ии|ия)?")),
    }


def _load_cache() -> dict[str, dict]:
    if not LIVE_CACHE_PATH.exists():
        return {}
    try:
        payload = json.loads(LIVE_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _save_cache(cache: dict[str, dict]) -> None:
    LIVE_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    LIVE_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_targets(*, handles: list[str], refresh_all: bool, limit: int) -> tuple[object, list[dict]]:
    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook["Brands"]
    headers = [cell.value for cell in worksheet[1]]
    header_map = {str(value): index + 1 for index, value in enumerate(headers)}

    if "Followers Count" not in header_map:
        worksheet.cell(row=1, column=worksheet.max_column + 1).value = "Followers Count"
        header_map["Followers Count"] = worksheet.max_column
    if "Posts Count" not in header_map:
        worksheet.cell(row=1, column=worksheet.max_column + 1).value = "Posts Count"
        header_map["Posts Count"] = worksheet.max_column

    selected: list[dict] = []
    wanted = {canonical_handle(handle) for handle in handles if canonical_handle(handle)}
    for row_idx in range(2, worksheet.max_row + 1):
        handle = str(worksheet.cell(row=row_idx, column=header_map["Handle"]).value or "").strip()
        profile_url = str(worksheet.cell(row=row_idx, column=header_map["Profile URL"]).value or "").strip()
        canonical = canonical_handle(handle)
        if not canonical or not profile_url:
            continue
        followers = int(worksheet.cell(row=row_idx, column=header_map["Followers Count"]).value or 0)
        if wanted and canonical not in wanted:
            continue
        if not refresh_all and followers > 0:
            continue
        selected.append(
            {
                "row_idx": row_idx,
                "handle": handle,
                "canonical_handle": canonical,
                "profile_url": profile_url,
            }
        )
        if limit and len(selected) >= limit:
            break
    return workbook, selected


async def _refresh_targets(targets: list[dict], *, profile_no: str) -> tuple[list[dict], Path]:
    artifacts, logger = setup_run_artifacts(PROJECT_ROOT, "instagram_common_followers_live")
    settings = AdsPowerSettings.from_project_root(PROJECT_ROOT)
    client = AdsPowerClient(
        AdsPowerSettings(
            base_url=settings.base_url,
            api_key=settings.api_key,
            profile_no=profile_no or settings.profile_no,
        )
    )
    started = client.start_profile(profile_no=profile_no or settings.profile_no, last_opened_tabs=False)

    playwright = await async_playwright().start()
    browser = None
    results: list[dict] = []
    try:
        browser, context, page = await connect_profile(playwright, started.ws_puppeteer, logger)
        await page.bring_to_front()
        for extra in list(context.pages)[1:]:
            try:
                await extra.close()
            except Exception:
                pass

        human: Humanizer = build_humanizer(page, {"humanization_policy": {}})

        for index, item in enumerate(targets, start=1):
            handle = item["handle"]
            profile_url = item["profile_url"]
            logger.info("Refreshing Instagram metrics %s/%s -> @%s", index, len(targets), handle)
            try:
                await page.goto(profile_url, wait_until="domcontentloaded", timeout=60000)
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except PlaywrightTimeoutError:
                    logger.info("Profile did not reach networkidle: %s", profile_url)
                await human.pause(1000, 1800)
                await dismiss_instagram_popups(page, human, logger)
                overview = await read_profile_overview(page)
                counts = _extract_profile_counts(
                    str(overview.get("followers_text") or ""),
                    str(overview.get("header_text") or ""),
                )
                if counts["followers"] <= 0:
                    counts["followers"] = int(overview.get("followers_count") or 0)
                results.append(
                    {
                        **item,
                        **counts,
                        "raw_followers_text": str(overview.get("followers_text") or ""),
                        "fetched_at": time_now_iso(),
                        "ok": True,
                    }
                )
                await human.pause(900, 1500)
            except Exception as exc:
                shot_path = artifacts.screenshots_dir / f"{canonical_handle(handle)}.png"
                try:
                    await capture_screenshot(page, shot_path, logger)
                except Exception:
                    pass
                results.append(
                    {
                        **item,
                        "followers": 0,
                        "following": 0,
                        "posts": 0,
                        "raw_followers_text": "",
                        "fetched_at": time_now_iso(),
                        "ok": False,
                        "error": str(exc),
                        "screenshot": str(shot_path),
                    }
                )
                await human.pause(800, 1400)
    finally:
        if browser is not None:
            try:
                await browser.close()
            except Exception:
                pass
        try:
            await playwright.stop()
        except Exception:
            pass
        try:
            client.stop_profile(profile_no=profile_no or settings.profile_no)
        except Exception:
            pass
    return results, artifacts.run_dir


def time_now_iso() -> str:
    from datetime import datetime, timezone

    return datetime.now(timezone.utc).isoformat()


def _apply_results(workbook, results: list[dict]) -> None:
    worksheet = workbook["Brands"]
    headers = [cell.value for cell in worksheet[1]]
    header_map = {str(value): index + 1 for index, value in enumerate(headers)}
    cache = _load_cache()

    for item in results:
        if not item.get("ok"):
            continue
        row_idx = int(item["row_idx"])
        worksheet.cell(row=row_idx, column=header_map["Followers Count"]).value = int(item.get("followers") or 0)
        worksheet.cell(row=row_idx, column=header_map["Posts Count"]).value = int(item.get("posts") or 0)
        cache[item["canonical_handle"]] = {
            "followers": int(item.get("followers") or 0),
            "following": int(item.get("following") or 0),
            "posts": int(item.get("posts") or 0),
            "profile_url": item.get("profile_url") or "",
            "raw_followers_text": item.get("raw_followers_text") or "",
            "fetched_at": item.get("fetched_at") or "",
        }

    workbook.save(WORKBOOK_PATH)
    _save_cache(cache)


def main() -> None:
    args = parse_args()
    workbook, targets = _load_targets(handles=args.handles, refresh_all=args.all, limit=args.limit)
    if not targets:
        print("No Instagram profiles matched the refresh criteria.")
        return

    results, run_dir = asyncio.run(_refresh_targets(targets, profile_no=args.profile_no.strip()))
    _apply_results(workbook, results)

    summary = {
        "workbook_path": str(WORKBOOK_PATH),
        "cache_path": str(LIVE_CACHE_PATH),
        "run_dir": str(run_dir),
        "targets": len(targets),
        "ok": sum(1 for item in results if item.get("ok")),
        "failed": sum(1 for item in results if not item.get("ok")),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
